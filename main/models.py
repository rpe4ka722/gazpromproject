from datetime import datetime

from django.contrib.auth.models import User
from django.db import models
from django.core.validators import MinValueValidator, MaxValueValidator, FileExtensionValidator
from django.shortcuts import reverse
from openpyxl import load_workbook


class RrlLine(models.Model):
    rrl_line_name: str = models.CharField(max_length=50, verbose_name='наименование линии')
    station_count = models.IntegerField(verbose_name='количество станций', blank=True, null=True)
    bandwidth = models.CharField(max_length=20, verbose_name='полоса пропускания', blank=True, null=True)

    @property
    def get_station_count_field(self):
        if self.station_count:
            return self.station_count
        else:
            return ""

    @property
    def get_bandwidth_field(self):
        if self.bandwidth:
            return self.bandwidth
        else:
            return ""

    def __str__(self):
        return self.rrl_line_name


class Position(models.Model):
    latitude_degrees = models.IntegerField(verbose_name='°',
                                           validators=[MinValueValidator(0, message='Значение градусов не может '
                                                                                    'быть отрицательным'),
                                                       MaxValueValidator(180, message='Значение градусов не может '
                                                                                      'быть больше 180')])

    latitude_minutes = models.IntegerField(verbose_name="'",
                                           validators=[MinValueValidator(0, message='Значение минут не может '
                                                                                    'быть отрицательным'),
                                                       MaxValueValidator(60, message='Значение минут не может '
                                                                                     'быть больше 60')])
    latitude_seconds = models.FloatField(verbose_name='"',
                                         validators=[MinValueValidator(0, message='Значение секунд не может '
                                                                                  'быть отрицательным'),
                                                     MaxValueValidator(60, message='Значение секунд не может '
                                                                                   'быть больше 60')])
    longitude_degrees = models.IntegerField(verbose_name='°',
                                            validators=[MinValueValidator(0, message='Значение градусов не может '
                                                                                     'быть отрицательным'),
                                                        MaxValueValidator(180, message='Значение градусов не может '
                                                                                       'быть больше 180')])
    longitude_minutes = models.IntegerField(verbose_name="'",
                                            validators=[MinValueValidator(0, message='Значение минут не может '
                                                                                     'быть отрицательным'),
                                                        MaxValueValidator(60, message='Значение минут не может '
                                                                                      'быть больше 60')])
    longitude_seconds = models.FloatField(verbose_name='"',
                                          validators=[MinValueValidator(0, message='Значение секунд не может '
                                                                                   'быть отрицательным'),
                                                      MaxValueValidator(60, message='Значение секунд не может '
                                                                                    'быть больше 60')])
    address = models.CharField(max_length=50, verbose_name='Адрес', blank=True, null=True)
    district = models.CharField(max_length=50, verbose_name='Район', blank=True, null=True)


class Department(models.Model):
    UCHASTOK_CHOICES = [
        ('Участок связи при заполярной промплощадке', 'Участок связи при заполярной промплощадке'),
        ('Участок связи №1', 'Участок связи №1'),
        ('Участок связи №2', 'Участок связи №2'),
        ('Участок связи №3', 'Участок связи №3'),
        ('Участок связи №4', 'Участок связи №4'),
        ('Участок связи №5', 'Участок связи №5'),
        ('Участок связи №6', 'Участок связи №6'),
        ('Участок связи №7', 'Участок связи №7'),
        ('Участок связи №8', 'Участок связи №8'),
        ('Участок связи №9', 'Участок связи №9'),
        ('Участок связи №10', 'Участок связи №10'),
        ('Участок связи №11', 'Участок связи №11'),
        ('Участок связи №12', 'Участок связи №12'),
        ('Участок связи №13', 'Участок связи №13'),
        ('Участок связи №14', 'Участок связи №14'),
        ('Участок связи №15', 'Участок связи №15'),
        ('Участок связи №16', 'Участок связи №16'),
        ('Участок связи №17', 'Участок связи №17'),
        ('Участок связи №18', 'Участок связи №18'),
        ('Участок связи №19', 'Участок связи №19'),
        ('Участок связи №20', 'Участок связи №20'),
        ('АПГ', 'АПГ'),
        ('Руководство', 'Руководство'),
        ('ПТО', 'ПТО'),
        ('ООТиПБ', 'ООТиПБ'),
        ('СЭСПД', 'СЭСПД'),
        ('группа РРЛ', 'группа РРЛ'),
        ('группа ЭПУС', 'группа ЭПУС'),
        ('группа РС', 'группа РС'),
        ('группа КС', 'группа КС'),
        ('группа ПД', 'группа ПД'),
    ]

    CEH_CHOICES = [
        ('Ноябрьский цех связи', 'НЦС'),
        ('Сургутский цех связи', 'СЦС'),
        ('Тюменский цех связи', 'ТЦС'),
        ('ПЛС', 'ПЛС'),
        ('Администрация', 'Администрация')
    ]

    uchastok = models.CharField(max_length=50, choices=UCHASTOK_CHOICES, verbose_name='участок связи')
    ceh = models.CharField(max_length=30, choices=CEH_CHOICES, verbose_name='цех', blank=True)
    is_prod = models.BooleanField(default=False)

    class Meta:
        ordering = ['ceh', 'uchastok']

    def __str__(self):
        return self.uchastok


class Object(models.Model):
    object_name = models.CharField(max_length=30, verbose_name='Наименование станции', unique=True)
    position = models.OneToOneField(Position, on_delete=models.PROTECT, verbose_name='Место расположения станции')
    rrl_line = models.ForeignKey(RrlLine, on_delete=models.DO_NOTHING, null=True, blank=True,
                                 verbose_name='Наименование линии связи')
    uchastok = models.ForeignKey(Department, on_delete=models.DO_NOTHING, blank=True, null=True,
                                 verbose_name='Участок связи')
    last_modify = models.ForeignKey(User, on_delete=models.DO_NOTHING)
    time_modify = models.DateTimeField(auto_now=True)


    def get_absolute_url(self):
        return reverse('main:object_detail', kwargs={'name': self.id})

    def __str__(self):
        return self.object_name

    def is_rich(self):
        if self.res.filter(related_rich__doc__isnull=False).exists():
            return True
        else:
            return False


class UploadedData(models.Model):
    file = models.FileField(upload_to='temp/', validators=[FileExtensionValidator(allowed_extensions=['xls', 'xlsx'],
                                                                                  message='Тип файла выбран неверно. '
                                                                                          'Выберите файл таблицы excel.'
                                                                                  )])

    def sheets(self):
        wb = load_workbook(self.file)
        return wb.sheetnames

    def headers(self, sheet, row):
        x = str(sheet)
        wb = load_workbook(self.file)
        active_data = wb[x]
        headers = []
        for col in active_data.iter_cols(min_row=row, max_row=row):
            for cell in col:
                headers.append(cell.value)
        return headers


class Header(models.Model):
    header = models.CharField(max_length=50)

    def __str__(self):
        return self.header


class Sheet(models.Model):
    sheet_name = models.CharField(max_length=50)

    def __str__(self):
        return self.sheet_name


class Choice(models.Model):
    sheet = models.ForeignKey(Sheet, verbose_name='Выберите вкладку для импорта', related_name='sheet',
                              on_delete=models.CASCADE)
    row = models.IntegerField(validators=[MinValueValidator(1, message='Значение не может быть меньше единицы')],
                              verbose_name='Для импорта выберите номер строки с заголовками таблицы')
    object_name = models.ForeignKey(Header, verbose_name='Выберите источник данных для названий объектов',
                                    related_name='object_name', on_delete=models.CASCADE, null=True)
    coords_lat = models.ForeignKey(Header, verbose_name='Выберите источник данных для широты',
                                   related_name='latitude', on_delete=models.CASCADE, null=True)
    coords_lon = models.ForeignKey(Header, verbose_name='Выберите источник данных для долготы',
                                   related_name='longitude', on_delete=models.CASCADE, null=True)
    district = models.ForeignKey(Header, verbose_name='Выберите источник данных для указания района',
                                 related_name='district', on_delete=models.CASCADE, null=True)
    address = models.ForeignKey(Header, verbose_name='Выберите источник данных для адресов объектов',
                                related_name='address', on_delete=models.CASCADE, null=True)


class Ozp(models.Model):
    object_name = models.ForeignKey(Object, on_delete=models.CASCADE, related_name='ozp')
    zamechanie_ozp = models.TextField(max_length=500)
    normative_documentation = models.CharField(max_length=100, blank=True, null=True)
    creation_date = models.DateField(auto_now_add=True)
    control_date = models.DateField(blank=True, null=True)
    zakrytie_date = models.DateField(blank=True, null=True)
    last_modify = models.ForeignKey(User, on_delete=models.DO_NOTHING)
    time_modify = models.DateTimeField(auto_now=True)
    is_done = models.BooleanField(default=False)

    def get_absolute_url(self):
        return reverse('main:ozp_details', kwargs={'ozp_id': self.id})


class Foto_zamechanya(models.Model):
    zamechanie = models.ForeignKey(Ozp, on_delete=models.CASCADE, related_name='foto_zamechania')
    foto = models.ImageField(upload_to='img/', validators=[FileExtensionValidator(allowed_extensions=['.png', '.jpeg',
                                                                                                      '.tiff', '.bmp',
                                                                                                      '.gif', 'jpg'],
                                                                                  message='Тип файла выбран неверно. '
                                                                                          'Выберите файл изображения.')]
                             )


class Foto_vipolnenya(models.Model):
    zamechanie = models.ForeignKey(Ozp, on_delete=models.CASCADE)
    foto = models.ImageField(upload_to='img/', validators=[FileExtensionValidator(allowed_extensions=['.png',
                                                                                                      '.jpg',
                                                                                                      '.jpeg',
                                                                                                      '.tiff',
                                                                                                      '.bmp',
                                                                                                      '.gif'],
                                                                                  message='Тип файла выбран '
                                                                                          'неверно. Выберите файл'
                                                                                          ' изображения.'
                                                                                  )])


class Podano_na_vipolnenie(models.Model):
    zamechanie = models.OneToOneField(Ozp, on_delete=models.CASCADE, related_name='podano_na_vipolnenie')
    podano = models.BooleanField(default=False)
    user = models.ForeignKey(User, on_delete=models.DO_NOTHING)
    time_podano = models.DateTimeField(auto_now=True)
    comment = models.TextField(max_length=500)
    otklonit_comment = models.TextField(max_length=500, blank=True, null=True)


class Technicheskie_usloviya(models.Model):
    name = models.CharField(max_length=50, verbose_name='Укажите наименование документа')
    organization = models.CharField(max_length=100, verbose_name='Укажите организацию которой выданы ТУ')
    proekt = models.CharField(max_length=150, verbose_name='Укажите наименование проекта')
    object = models.ForeignKey(Object, on_delete=models.DO_NOTHING, verbose_name='Выберите объект технических условий',
                               related_name='tu')
    doc = models.FileField(upload_to='media/', validators=[FileExtensionValidator(allowed_extensions=['png', 'jpeg',
                                                                                                      'tiff', 'bmp',
                                                                                                      'gif', 'pdf',
                                                                                                      'doc', 'docx',
                                                                                                      'jpg'],
                                                                                  message='Тип файла выбран неверно.')],
                           verbose_name='Загрузите копию технических условий')
    date = models.DateField(verbose_name='Укажите дату технических условий')
    description = models.TextField(max_length=500, verbose_name='Напишите краткое описание технических условий')
    expire_date = models.DateField(verbose_name='Укажите дату окончания действия технических условий')

    def __str__(self):
        return self.name

    def is_active(self):
        if self.expire_date > datetime.now().date():
            return True
        else:
            return False


class Technicheskaya_documentacia(models.Model):
    name = models.CharField(max_length=50, verbose_name='Укажите наименование документа')
    object = models.ForeignKey(Object, on_delete=models.DO_NOTHING, verbose_name='Выберите объект',
                               related_name='tehdoc')
    doc = models.FileField(upload_to='media/', validators=[FileExtensionValidator(allowed_extensions=['png', 'jpeg',
                                                                                                      'tiff', 'bmp',
                                                                                                      'gif', 'pdf',
                                                                                                      'doc', 'docx',
                                                                                                      'jpg'],
                                                                                  message='Тип файла выбран неверно.')],
                           verbose_name='Загрузите документ')
    date = models.DateField(verbose_name='Укажите дату документа')
    description = models.TextField(max_length=500, verbose_name='Напишите краткое описание документа')

    def __str__(self):
        return self.name


class Scheme(models.Model):
    name = models.CharField(max_length=50, verbose_name='Укажите наименование схемы')
    object = models.ForeignKey(Object, on_delete=models.DO_NOTHING, verbose_name='Выберите объект',
                               related_name='scheme')
    doc = models.FileField(upload_to='media/', validators=[FileExtensionValidator(allowed_extensions=['png', 'jpeg',
                                                                                                      'tiff', 'bmp',
                                                                                                      'gif', 'pdf',
                                                                                                      'doc', 'docx',
                                                                                                      'jpg', 'vsd',
                                                                                                      'vsdx', 'vdx'],
                                                                                  message='Тип файла выбран неверно.')],
                           verbose_name='Загрузите схему')
    date = models.DateField(verbose_name='Укажите дату документа')

    def __str__(self):
        return self.name


class Object_Foto(models.Model):
    object = models.ForeignKey(Object, on_delete=models.CASCADE, related_name='foto',
                               verbose_name='Выберите объект изображенный на фотографии')
    foto = models.ImageField(upload_to='media', validators=[FileExtensionValidator(message='Тип файла выбран неверно.',
                                                                                   allowed_extensions=['png', 'jpeg',
                                                                                                       'tiff', 'bmp',
                                                                                                       'gif', 'jpg'])],
                             verbose_name='Выберите фотографию АМС')
    year = models.DateField(default=None, verbose_name='Укажите ориентировочную дату фото')
