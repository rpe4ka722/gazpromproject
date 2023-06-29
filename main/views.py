from datetime import datetime, timedelta

from django.contrib.auth.models import User
from django.core.exceptions import ObjectDoesNotExist
from openpyxl.styles.numbers import BUILTIN_FORMATS

from account.forms import UserCreationForm
from account.models import Userprofile
from main.scripts import is_staff, form_errors_text, object_filter
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponseNotFound, HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse

from rich.models import Rich, Res
from .forms import ObjectForm, RrlForm, PositionForm, UploadXlsForm, UploadChoiceForm, ImportForm, FilterForm, \
    Tu_Create_Form, Tehdoc_Create_Form, Scheme_Create_Form, Object_Name_Form, Object_Uchastok_Form, Coordinat_Form, \
    Object_Address_Form, Object_District_Form, Object_Foto_Form, List_Foto_Form
from .models import Object, Position, RrlLine, Sheet, Header, UploadedData, Choice, Department, Ozp, \
    Foto_zamechanya, Foto_vipolnenya, Podano_na_vipolnenie, Technicheskie_usloviya, Technicheskaya_documentacia, \
    Scheme, Object_Foto
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Border, Side, Alignment, Font

from .scripts import str_to_coord


@login_required(login_url='account:login')
def index(request):
    return render(request, 'main/templates/index.html')


@login_required(login_url='account:login')
def object_index(request):
    return render(request, 'main/templates/object_index.html')


@login_required(login_url='account:login')
@is_staff
def create_object(request):
    obj_form = ObjectForm(request.POST or None)
    pos_form = PositionForm(request.POST or None)
    if request.method == 'POST':
        if obj_form.is_valid() and pos_form.is_valid():
            cd_obj = obj_form.cleaned_data
            cd_pos = pos_form.cleaned_data
            pos = Position(latitude_degrees=cd_pos['latitude_degrees'],
                           latitude_minutes=cd_pos['latitude_minutes'], latitude_seconds=cd_pos['latitude_seconds'],
                           longitude_degrees=cd_pos['longitude_degrees'], longitude_minutes=cd_pos['longitude_minutes'],
                           longitude_seconds=cd_pos['longitude_seconds'], address=cd_pos['address'],
                           district=cd_pos['district'])
            pos.save()
            user = request.user
            time = datetime.now()
            obj = Object(object_name=cd_obj['object_name'], position=pos, rrl_line=cd_obj['rrl_line'],
                         uchastok=cd_obj['uchastok'], last_modify=user, time_modify=time)
            obj.save()
            return redirect('main:object_detail', obj.id)
        else:
            msg = 'Веденные данные некорректны'
            return render(request, "main/templates/object_creation.html", {'obj_form': obj_form,
                                                                           'pos_form': pos_form, 'msg': msg})
    else:
        obj_form = ObjectForm()
        pos_form = PositionForm()
    return render(request, "main/templates/object_creation.html", {'obj_form': obj_form, 'pos_form': pos_form})


@login_required(login_url='account:login')
@is_staff
def create_rrl(request):
    if request.method == 'POST':
        form = RrlForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            rrl = RrlLine(rrl_line_name=cd['rrl_line_name'], station_count=cd['station_count'],
                          bandwidth=cd['bandwidth'])
            rrl.save()
            return redirect('main:rrl_line')
        else:
            msg = 'Вы ввели некорректные данные'
            context = {'form': form, 'msg': msg}
            return render(request, "main/templates/rrl_creation.html", context)
    else:
        form = RrlForm()
    return render(request, "main/templates/rrl_creation.html", {'form': form})


@login_required(login_url='account:login')
def objects_list(request, msg=''):
    filter_form = FilterForm()
    objectslist = Object.objects.select_related('uchastok', 'position').prefetch_related('ams', 'ozp', 'res', 'tehdoc',
                                                                                         'tu', 'scheme', 'foto').all
    if request.method == 'GET':
        try:
            data = FilterForm(request.GET or None)
            objects, filtered_by = object_filter(data)
            filtered_by_str = ' '.join(filtered_by)
            objectslist = objects.select_related('uchastok', 'position').prefetch_related('ams', 'ozp', 'res', 'tehdoc',
                                                                                          'tu', 'scheme', 'foto')
            context = {'objectlist': objectslist, 'msg': msg, 'filter_form': filter_form,
                       'filtered_by': filtered_by_str}
            return render(request, 'main/templates/objects.html', context)
        except (ValueError, TypeError):
            filtered_by = ''
            context = {'objectlist': objectslist, 'msg': msg, 'filter_form': filter_form,
                       'filtered_by': filtered_by}
            return render(request, 'main/templates/objects.html', context)


@login_required(login_url='account:login')
def object_detail(request, name: int, msg=''):
    selected_object = Object.objects.get(id=name)
    selected_rich = Rich.objects.filter(related_res__in=Res.objects.filter(related_object=selected_object)).distinct()
    name_form = Object_Name_Form()
    uchastok_form = Object_Uchastok_Form()
    coord_form = Coordinat_Form()
    address_form = Object_Address_Form()
    district_form = Object_District_Form()
    foto_form = Object_Foto_Form()
    context = {'i': selected_object, 'rich': selected_rich, 'name_form': name_form, 'msg': msg,
               'uchastok_form': uchastok_form, 'coord_form': coord_form, 'address_form': address_form,
               'district_form': district_form, 'foto_form': foto_form}
    return render(request, 'main/templates/object_detail.html', context)


@login_required(login_url='account:login')
@is_staff
def upload_xls(request):
    if request.method == 'POST':
        form = UploadXlsForm(request.POST, request.FILES)
        if form.is_valid():
            file = form.save()
            object_id = file.id
            sheets = file.sheets()
            for i in sheets:
                inst = Sheet(sheet_name=i)
                inst.save()
            return redirect(reverse('main:upload_choice', kwargs={'object_id': object_id}))
    else:
        form = UploadXlsForm()
    return render(request, "main/templates/upload.html", {'form': form})


@login_required(login_url='account:login')
@is_staff
def upload_choice(request, object_id):
    if request.method == 'POST':
        form = UploadChoiceForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            instance = Choice(sheet=cd['sheet'], row=cd['row'])
            instance.save()
            sheet = instance.sheet
            row = instance.row
            data = UploadedData.objects.get(id=object_id)
            headers_list = data.headers(sheet, row)
            for i in headers_list:
                x = Header(header=i)
                x.save()
            return redirect(reverse('main:import_data', kwargs={'choice': instance.id, 'object_id': object_id}))
    else:
        form = UploadChoiceForm()
    return render(request, "main/templates/upload_1.html", {'form': form, 'object_id': object_id})


@login_required(login_url='account:login')
@is_staff
def import_data(request, choice, object_id):
    if request.method == 'POST':
        form = ImportForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            object_name = Header.objects.get(header=cd['object_name'])
            coords_lat = Header.objects.get(header=cd['coords_lat'])
            coords_lon = Header.objects.get(header=cd['coords_lon'])
            district = Header.objects.get(header=cd['district'])
            address = Header.objects.get(header=cd['address'])
            choice = Choice.objects.get(id=choice)
            choice.object_name = object_name
            choice.coords_lat = coords_lat
            choice.coords_lon = coords_lon
            choice.district = district
            choice.address = address
            choice.save()
            data = UploadedData.objects.get(id=object_id)
            wb = load_workbook(filename=data.file)
            sheet_name = str(choice.sheet.sheet_name)
            ws = wb[sheet_name]
            row_count = ws.max_row
            x: dict = {}
            for cols in ws.iter_cols(min_row=choice.row, max_row=choice.row):
                for cell in cols:
                    if str(cell.value) == str(choice.object_name):
                        x.update({'object_name': cell.column})
                    if str(cell.value) == str(choice.coords_lat):
                        x.update({'coords_lat': cell.column})
                    if str(cell.value) == str(choice.coords_lon):
                        x.update({'coords_lon': cell.column})
                    if str(cell.value) == str(choice.district):
                        x.update({'district': cell.column})
                    if str(cell.value) == str(choice.address):
                        x.update({'address': cell.column})
            for row in range(row_count):
                lat = ws.cell(row=row + 5, column=x['coords_lat']).value
                lon = ws.cell(row=row + 5, column=x['coords_lon']).value
                if lat is None or lon is None:
                    pass
                else:
                    y = str_to_coord(lat)
                    z = str_to_coord(lon)
                    pos = Position(latitude_degrees=y[0], latitude_minutes=y[1], latitude_seconds=y[2],
                                   longitude_degrees=z[0], longitude_minutes=z[1], longitude_seconds=z[2],
                                   address=ws.cell(row=row + 5, column=x['address']).value,
                                   district=ws.cell(row=row + 5, column=x['district']).value)
                    print(pos.address)
                    pos.save()
                    user = request.user
                    time = datetime.now()
                    obj = Object(object_name=ws.cell(row=row + 5, column=x['object_name']).value, position=pos,
                                 last_modify=user, time_modify=time)
                    obj.save()
            for i in Header.objects.all():
                i.delete()
            for i in Choice.objects.all():
                i.delete()
            for i in Sheet.objects.all():
                i.delete()
            for i in UploadedData.objects.all():
                i.delete()
            return redirect('main:objects')

    else:
        form = ImportForm()
    return render(request, "main/templates/import.html", {'form': form, 'choice': choice, 'object_id': object_id})


@login_required(login_url='account:login')
def structure(request, msg=''):
    if request.GET:
        objectslist = User.objects.all()
        if request.GET['department'] == 'all':
            filtered_by = 'Отсутствует'
        else:
            dep = request.GET['department']
            filtered_by = Department.objects.get(id=dep)
            objectslist = User.objects.filter(profile__in=Userprofile.objects.filter(department=dep))
    else:
        objectslist = User.objects.all()
        filtered_by = 'all'
    department = Department.objects.all()
    context = {'objectlist': objectslist, 'department': department, 'filtered_by': filtered_by, 'msg': msg}
    return render(request, "main/templates/structure.html", context)


@login_required(login_url='account:login')
@is_staff
def user_delete(request, user_id):
    try:
        current_user = request.user
        user = User.objects.get(id=user_id)
        if user == current_user:
            msg = 'Вы не можете удалить пользователя под которым авторизованы'
            return structure(request, msg=msg)
        else:
            user.delete()
    except ObjectDoesNotExist:
        pass
    return redirect('main:structure')


@login_required(login_url='account:login')
@is_staff
def user_create(request, msg=''):
    form = UserCreationForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            cd = form.cleaned_data
            if User.objects.filter(username=cd['username']).exists():
                msg = "Такой пользователь уже зарегестирован"
            else:
                user = User.objects.create_user(username=cd['username'], password=cd['password'], email=cd['email'],
                                                first_name=cd['first_name'], last_name=cd['last_name'],
                                                is_staff=cd['is_staff'])
                user.save()
                department = request.POST['department']
                profile = Userprofile.objects.create(user=user, department=Department.objects.get(id=department))
                profile.save()
                return redirect('main:structure')
        else:
            msg = form_errors_text(form)
    return render(request, "main/templates/user_creation.html", {'form': form, 'msg': msg})


@login_required(login_url='account:login')
def export_xls_department(request, ceh):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="dep.xlsx"'
    wb = Workbook()
    ws = wb.active
    row_num = 1
    columns = ['Цех', 'Участок', 'Начальник участка']
    font = Font(name='TimesNewRoman', sz=11, bold=True)
    font2 = Font(name='TimesNewRoman', sz=11, bold=False)
    border = Border(left=Side(border_style='thin', color='FF000000'),
                    right=Side(border_style='thin', color='FF000000'),
                    top=Side(border_style='thin', color='FF000000'),
                    bottom=Side(border_style='thin', color='FF000000'))
    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=True,
                          shrink_to_fit=False,
                          indent=0)
    for col_num in range(len(columns)):
        ws.cell(row_num, col_num + 1, columns[col_num])
        ws.cell(row_num, col_num + 1).font = font
        ws.cell(row_num, col_num + 1).border = border
        ws.cell(row_num, col_num + 1).alignment = alignment
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 30
    if ceh == 'all':
        rows = Department.objects.all()
    elif ceh == 'Отсутствует':
        rows = Department.objects.filter(ceh__exact='')
    else:
        rows = Department.objects.filter(ceh__exact=ceh)
    for row in range(rows.count()):
        atr_list = [rows[row].ceh, rows[row].uchastok, rows[row].employee.employee_name + ' ' +
                    rows[row].employee.employee_last_name]
        for col_num in range(len(atr_list)):
            ws.cell(row + 2, col_num + 1, atr_list[col_num])
            ws.cell(row + 2, col_num + 1).font = font2
            ws.cell(row + 2, col_num + 1).border = border
            ws.cell(row + 2, col_num + 1).alignment = alignment
    wb.save(response)
    return response


@login_required(login_url='account:login')
def rrl_line(request):
    objectslist = RrlLine.objects.all()
    context = {'objectlist': objectslist}
    return render(request, "main/templates/rrl.html", context)


@login_required(login_url='account:login')
@is_staff
def delete_rrl(request, pk):
    try:
        obj = RrlLine.objects.get(id=pk)
        obj.delete()
        return redirect('main:rrl_line')
    except Department.DoesNotExist:
        return HttpResponseNotFound("<h2>Not found</h2>")


@login_required(login_url='account:login')
@is_staff
def change_rrl(request, pk):
    obj = RrlLine.objects.get(id=pk)
    if request.method == 'POST':
        form = RrlForm(request.POST)
        if form.is_valid():
            cd = form.cleaned_data
            obj.rrl_line_name = cd['rrl_line_name']
            obj.station_count = cd['station_count']
            obj.bandwidth = cd['bandwidth']
            obj.save()
            return redirect('main:rrl_line')
        else:
            msg = 'Вы ввели некорректные данные'
            context = {'form': form, 'msg': msg, 'pk': pk, 'obj': obj}
            return render(request, "main/templates/change_rrl.html", context)
    else:
        form = RrlForm()
        context = {'form': form, 'pk': pk, 'obj': obj}
    return render(request, "main/templates/change_rrl.html", context)


@login_required(login_url='account:login')
def export_objects(request):
    if request.method == 'GET':
        try:
            data = FilterForm(request.GET or None)
            objects, filtered_by = object_filter(data)
        except (ValueError, TypeError):
            return redirect('main:objects', msg='Неизвестная ошибка')
        # выгрузка в xls -----------------------------------------------------
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="objects.xlsx"'
        wb = Workbook()
        rows = objects
        ws = wb.active
        row_num = 1
        columns = ['Наименование объекта', 'Цех', 'Участок', 'Координаты объекта', 'Адрес', 'Район размещения объекта',
                   'Магистральная линия']
        font = Font(name='TimesNewRoman', sz=11, bold=True)
        font2 = Font(name='TimesNewRoman', sz=11, bold=False)
        border = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'))
        alignment = Alignment(horizontal='center',
                              vertical='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
        for col_num in range(len(columns)):
            ws.cell(row_num, col_num + 1, columns[col_num])
            ws.cell(row_num, col_num + 1).font = font
            ws.cell(row_num, col_num + 1).border = border
            ws.cell(row_num, col_num + 1).alignment = alignment
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        for row in range(rows.count()):
            if rows[row].rrl_line is None:
                atr_list = [rows[row].object_name, rows[row].uchastok.ceh, rows[row].uchastok.uchastok,
                            f'{rows[row].position.latitude_degrees}° {rows[row].position.latitude_minutes}\' '
                            f'{rows[row].position.latitude_seconds}" С.Ш. {rows[row].position.longitude_degrees}° '
                            f'{rows[row].position.longitude_minutes}\' {rows[row].position.longitude_seconds}" В.Д.',
                            rows[row].position.address, rows[row].position.district, '-']
            else:
                atr_list = [rows[row].object_name, rows[row].uchastok.ceh, rows[row].uchastok.uchastok,
                            f'{rows[row].position.latitude_degrees}° {rows[row].position.latitude_minutes}\' '
                            f'{rows[row].position.latitude_seconds}" С.Ш. {rows[row].position.longitude_degrees}° '
                            f'{rows[row].position.longitude_minutes}\' {rows[row].position.longitude_seconds}" В.Д.',
                            rows[row].position.address, rows[row].position.district, rows[row].rrl_line.rrl_line_name]
            ws.cell(row + 2, 7).number_format = BUILTIN_FORMATS[1]
            for col_num in range(len(atr_list)):
                ws.cell(row + 2, col_num + 1, atr_list[col_num])
                ws.cell(row + 2, col_num + 1).font = font2
                ws.cell(row + 2, col_num + 1).border = border
                ws.cell(row + 2, col_num + 1).alignment = alignment
        wb.save(response)
        return response


@login_required(login_url='account:login')
@is_staff
def change_object(request, pk):
    try:
        Object.objects.get(id=pk)
    except Department.DoesNotExist:
        return HttpResponseNotFound("<h2>Not found</h2>")
    if request.method == 'POST':
        obj_form = ObjectForm(request.POST)
        pos_form = PositionForm(request.POST)
        if obj_form.is_valid() and pos_form.is_valid():
            cd_obj = obj_form.cleaned_data
            cd_pos = pos_form.cleaned_data
            pos = Position(latitude_degrees=cd_pos['latitude_degrees'],
                           latitude_minutes=cd_pos['latitude_minutes'], latitude_seconds=cd_pos['latitude_seconds'],
                           longitude_degrees=cd_pos['longitude_degrees'], longitude_minutes=cd_pos['longitude_minutes'],
                           longitude_seconds=cd_pos['longitude_seconds'], address=cd_pos['address'],
                           district=cd_pos['district'])
            pos.save()
            user = request.user
            time = datetime.now()
            obj = Object(object_name=cd_obj['object_name'], position=pos, rrl_line=cd_obj['rrl_line'],
                         uchastok=cd_obj['uchastok'], last_modify=user, time_modify=time)
            obj.save()
            return redirect('/')
        else:
            msg = 'Введенные данные некорректны'
            return render(request, "main/templates/object_creation.html", {'obj_form': obj_form,
                                                                           'pos_form': pos_form, 'msg': msg})


@login_required(login_url='account:login')
@is_staff
def ozp_create(request):
    objects = Object.objects.all()
    if request.method == 'POST':
        uploaded_file = request.FILES['foto_do']
        accepted_list = ('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.gif')
        name = str(uploaded_file.name)
        associated_object = request.POST['associated_object']
        zamechanie_ozp = str(request.POST['zamechanie_ozp'])
        normative_documentation = str(request.POST['normative_documentation'])
        if name.endswith(accepted_list):
            x = Ozp.objects.create(object_name=Object.objects.get(object_name=associated_object),
                                   zamechanie_ozp=zamechanie_ozp, normative_documentation=normative_documentation,
                                   last_modify=request.user)
            x.save()
            y = Foto_zamechanya.objects.create(zamechanie=x, foto=uploaded_file)
            y.save()
            return redirect("main:ozp")
        else:
            msg = 'Загружать необходимо файл изображения!'
            context = {'objects': objects, 'msg': msg}
            return render(request, "main/templates/ozp_creation.html", context)
    context = {'objects': objects}
    return render(request, "main/templates/ozp_creation.html", context)


@login_required(login_url='account:login')
def ozp(request):
    # получение списков данных для меню выбора -------------------
    objectslist = Ozp.objects.all()
    filter_form = FilterForm()
    year_list = []
    for i in objectslist:
        try:
            year = i.zakrytie_date.year
            if str(year) in year_list:
                pass
            else:
                year_list.append(str(i.zakrytie_date.year))
        except AttributeError:
            pass
    # --------------------------------------------------------------
    # получение введенных в фильр данных --------------------------
    if request.method == 'GET':
        data = FilterForm(request.GET or None)
        objects, filtered_by = object_filter(data)
        vipolnenie = request.GET.get('vipolnenie')
        na_vipolnenie = request.GET.get('na_vipolnenie')
        this_month = request.GET.get('this_month')
    # ------------------------------------------------------------------------
    # получение списка выбранных годов----------------------------------------
        year_filter = []
        for i in year_list:
            q = ''.join(['year_', str(i)])
            try:
                value = request.GET.get(q)
                if value is None:
                    pass
                else:
                    year_filter.append(value)
            except KeyError:
                pass
        if not year_filter:
            year_filter = year_list
    # -------------------------------------------------------------------------
    # фильтрация по объектам сети----------------------------------------------
        filtered_ozp_objects = Ozp.objects.filter(object_name__in=objects)
    # фильтрация по выполнению -----------------------------------------------------
        if vipolnenie == '1':
            filtered_by.append('Выполненным')
            filtered_ozp_objects = filtered_ozp_objects.filter(is_done=True)
        elif vipolnenie == '0':
            filtered_by.append('Невыполненным')
            filtered_ozp_objects = filtered_ozp_objects.filter(is_done=False)
        else:
            pass
    # -----------------------------------------------------------------------------
    # фильтрация по поданным на выполннеине ----------------------------------------
        if na_vipolnenie == '1':
            filtered_by.append('Поданным на устранение')
            filtered_podano = Podano_na_vipolnenie.objects.filter(podano=True)
            filtered_ozp_objects = filtered_ozp_objects.filter(podano_na_vipolnenie__in=filtered_podano)
        elif na_vipolnenie == '0':
            filtered_by.append('Не поданным на устранение')
            filtered_podano = Podano_na_vipolnenie.objects.filter(podano=False)
            filtered_ozp_objects = filtered_ozp_objects.filter(podano_na_vipolnenie__in=filtered_podano)
        else:
            pass
    # -------------------------------------------------------------------
    # фильтрация по году-------------------------------------------------
        if year_filter != year_list:
            filtered_ozp_objects = filtered_ozp_objects.filter(zakrytie_date__year__in=year_filter)
            filtered_by_year = 'По следующим годам ' + ' '.join(year_filter)
            filtered_by.append(filtered_by_year)
    # -------------------------------------------------------------------
    # фильтрация по выполнению в течении месяца--------------------------
        if this_month == 'this_month':
            month = datetime.today() + timedelta(days=30)
            filtered_ozp_objects = filtered_ozp_objects.filter(zakrytie_date__lt=month)
            filtered_by.append('По необходимости выполнения в течении месяца')
    # -------------------------------------------------------------------
        filtered_by_str = ' '.join(filtered_by)
        context = {'objectlist': filtered_ozp_objects, 'year': year_list, 'filtered_by': filtered_by_str,
                   'filter_form': filter_form}
        return render(request, "main/templates/ozp.html", context)
    else:
        filtered_ozp_objects = Ozp.objects.all()
        filtered_by = str('')
        context = {'objectlist': filtered_ozp_objects, 'filter_form': filter_form, 'year': year_list,
                   'filtered_by': filtered_by}
        return render(request, "main/templates/ozp.html", context)


# функция получения отфильтрованных департаментов по цеху и участку-----------------------
@login_required(login_url='account:login')
def department_filter(ceh, uchastok):
    filtered_by = []
    if ceh == 'all' and uchastok == 'all':
        filtered_department = Department.objects.all()
    elif ceh == 'all' and uchastok != 'all':
        filtered_department = Department.objects.filter(uchastok=uchastok)
        filtered_by = [uchastok]
    elif ceh != 'all' and uchastok == 'all':
        filtered_department = Department.objects.filter(ceh=ceh)
        filtered_by = [ceh]
    else:
        filtered_department = Department.objects.filter(Q(ceh=ceh), Q(uchastok=uchastok))
        filtered_by = [uchastok, ceh]
    return filtered_department, filtered_by
# --------------------------------------------------------------------------------------------


@login_required(login_url='account:login')
def ozp_details(request, ozp_id):
    ozp_object = Ozp.objects.get(id=ozp_id)
    foto_do = Foto_zamechanya.objects.filter(zamechanie=ozp_object)
    posle_foto = Foto_vipolnenya.objects.filter(zamechanie=ozp_object)
    context = {'i': ozp_object, 'foto_do': foto_do, 'posle_foto': posle_foto}
    return render(request, "main/templates/ozp_details.html", context)


@login_required(login_url='account:login')
@is_staff
def ozp_zamechanie_change(request, ozp_id):
    obj = Ozp.objects.get(id=ozp_id)
    if request.method == 'POST':
        zamechanie = request.POST['text_zamechanya']
        obj.zamechanie_ozp = str(zamechanie)
        obj.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
@is_staff
def ozp_normative_change(request, ozp_id):
    obj = Ozp.objects.get(id=ozp_id)
    if request.method == 'POST':
        normative = request.POST['text_normative']
        obj.normative_documentation = str(normative)
        obj.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
@is_staff
def srok_ustranenia_change(request, ozp_id):
    obj = Ozp.objects.get(id=ozp_id)
    if request.method == 'POST':
        srok = request.POST['srok']
        if srok:
            obj.zakrytie_date = srok
            obj.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
@is_staff
def control_srok_change(request, ozp_id):
    obj = Ozp.objects.get(id=ozp_id)
    if request.method == 'POST':
        srok = request.POST['srok']
        if srok:
            obj.control_date = srok
            obj.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
@is_staff
def foto_zamechanie_add(request, ozp_id):
    obj = Ozp.objects.get(id=ozp_id)
    if request.method == 'POST':
        try:
            uploaded_file = request.FILES['foto']
            x = Foto_zamechanya.objects.create(zamechanie=obj, foto=uploaded_file)
            x.save()
        except KeyError:
            pass
        obj.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
@is_staff
def foto_do_delete(request, f_id):
    x = Foto_zamechanya.objects.get(id=f_id)
    ozp_id = x.zamechanie.id
    z = Ozp.objects.get(id=ozp_id)
    x.delete()
    z.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
@is_staff
def posle_foto_delete(request, f_id):
    x = Foto_vipolnenya.objects.get(id=f_id)
    ozp_id = x.zamechanie.id
    z = Ozp.objects.get(id=ozp_id)
    x.delete()
    z.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
def foto_vipolnenie_add(request, ozp_id):
    obj = Ozp.objects.get(id=ozp_id)
    if request.method == 'POST':
        try:
            uploaded_file = request.FILES['foto']
            x = Foto_vipolnenya.objects.create(zamechanie=obj, foto=uploaded_file)
            x.save()
        except KeyError:
            pass
        obj.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
@is_staff
def accept(request, ozp_id):
    obj = Ozp.objects.get(id=ozp_id)
    obj.is_done = True
    obj.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
def podano_na_vipolnenie(request, ozp_id):
    ozp_object = Ozp.objects.get(id=ozp_id)
    if request.method == 'POST':
        comment = request.POST['comment']
        user = request.user
        time = datetime.now()
        try:
            obj = Podano_na_vipolnenie.objects.get(zamechanie=ozp_object)
            try:
                uploaded_file = request.FILES['foto']
                x = Foto_vipolnenya.objects.create(zamechanie=ozp_object, foto=uploaded_file)
                x.save()
            except KeyError:
                pass
            try:
                obj.comment = comment
                obj.otklonit_comment = None
                obj.podano = True
                obj.user = user
                obj.time_podano = time
                obj.save()
            except KeyError:
                pass
        except:
            try:
                uploaded_file = request.FILES['foto']
                x = Foto_vipolnenya.objects.create(zamechanie=ozp_object, foto=uploaded_file)
                x.save()
            except KeyError:
                pass
            try:
                y = Podano_na_vipolnenie.objects.create(comment=comment, otklonit_comment=None, zamechanie=ozp_object,
                                                        podano=True, user=user, time_podano=time)
                y.save()
            except KeyError:
                pass
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
@is_staff
def not_accept(request, ozp_id):
    obj = Ozp.objects.get(id=ozp_id)
    y = Podano_na_vipolnenie.objects.get(zamechanie=obj)
    user = request.user
    time = datetime.now()
    if request.method == 'POST':
        otklonit_comment = request.POST['comment']
        y.otklonit_comment = otklonit_comment
        y.user = user
        y.time_podano = time
        y.podano = False
        y.save()
    return redirect('main:ozp_details', ozp_id=ozp_id)


@login_required(login_url='account:login')
def export_xls_ozp(request):
    if request.method == 'GET':
        try:
            ceh = request.GET['ceh']
            uchastok = request.GET['uchastok']
            vipolnenie = request.GET['vipolnenie']
            na_vipolnenie = request.GET['na_vipolnenie']
            this_month = request.GET['this_month']
            objectslist = Ozp.objects.all()
            year_list = []
            year_filter = []
            for i in objectslist:
                try:
                    year = i.zakrytie_date.year
                    if str(year) in year_list:
                        pass
                    else:
                        year_list.append(str(i.zakrytie_date.year))
                except AttributeError:
                    pass
            for i in year_list:
                q = ''.join(['year_', str(i)])
                try:
                    year_filter.append(request.GET[q])
                except KeyError:
                    pass
            # -------------------------------------------------------------------------
            # фильтрация по объектам сети----------------------------------------------
            filtered_departments = department_filter(ceh, uchastok)[0]
            filtered_objects = Object.objects.filter(uchastok__in=filtered_departments)
            filtered_ozp_objects = Ozp.objects.filter(object_name__in=filtered_objects)
            # фильтрация по выполнению -----------------------------------------------------
            if vipolnenie == '1':
                filtered_ozp_objects = filtered_ozp_objects.filter(is_done=True)
            elif vipolnenie == '0':
                filtered_ozp_objects = filtered_ozp_objects.filter(is_done=False)
            else:
                pass
            # -----------------------------------------------------------------------------
            # фильтрация по поданным на выполннеине ----------------------------------------
            if na_vipolnenie == '1':
                filtered_podano = Podano_na_vipolnenie.objects.filter(podano=True)
                filtered_ozp_objects = filtered_ozp_objects.filter(podano_na_vipolnenie__in=filtered_podano)
            elif na_vipolnenie == '0':
                filtered_podano = Podano_na_vipolnenie.objects.filter(podano=False)
                filtered_ozp_objects = filtered_ozp_objects.filter(podano_na_vipolnenie__in=filtered_podano)
            else:
                pass
            # -------------------------------------------------------------------
            # фильтрация по году-------------------------------------------------
            if year_filter != year_list:
                filtered_ozp_objects = filtered_ozp_objects.filter(zakrytie_date__year__in=year_filter)
            else:
                pass
            # -------------------------------------------------------------------
            # фильтрация по выполнению в течении месяца--------------------------
            if this_month == 'this_month':
                month = datetime.today() + timedelta(days=30)
                filtered_ozp_objects = filtered_ozp_objects.filter(zakrytie_date__lt=month)
            else:
                pass
        except KeyError:
            filtered_ozp_objects = Ozp.objects.all()
        # выгрузка в xls -----------------------------------------------------
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ozp.xlsx"'
        wb = Workbook()
        rows = filtered_ozp_objects
        ws = wb.active
        row_num = 1
        columns = ['Объект', 'Участок связи', 'Цех', 'Содержание замечания', 'Нормативная документация',
                   'Контрольная дата', 'Дата устранения', 'Устранено', 'Подано на устарнение']
        font = Font(name='TimesNewRoman', sz=11, bold=True)
        font2 = Font(name='TimesNewRoman', sz=11, bold=False)
        border = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'))
        alignment = Alignment(horizontal='center',
                              vertical='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
        for col_num in range(len(columns)):
            ws.cell(row_num, col_num + 1, columns[col_num])
            ws.cell(row_num, col_num + 1).font = font
            ws.cell(row_num, col_num + 1).border = border
            ws.cell(row_num, col_num + 1).alignment = alignment
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 30
        for row in range(rows.count()):
            if rows[row].is_done:
                x = 'Да'
            else:
                x = 'Нет'
            try:
                if rows[row].podano_na_vipolnenie.podano:
                    y = 'Да'
                else:
                    y = 'Нет'
            except:
                y = 'Нет'
            atr_list = [rows[row].object_name.object_name, rows[row].object_name.uchastok.ceh,
                        rows[row].object_name.uchastok.uchastok, rows[row].zamechanie_ozp,
                        rows[row].normative_documentation, rows[row].control_date, rows[row].zakrytie_date, x, y]
            for col_num in range(len(atr_list)):
                ws.cell(row + 2, col_num + 1, atr_list[col_num])
                ws.cell(row + 2, col_num + 1).font = font2
                ws.cell(row + 2, col_num + 1).border = border
                ws.cell(row + 2, col_num + 1).alignment = alignment
        wb.save(response)
        return response


@login_required(login_url='account:login')
@is_staff
def delete_ozp(request, ozp_id):
    try:
        ozp_object = Ozp.objects.get(id=ozp_id)
        ozp_object.delete()
        return redirect('main:ozp')
    except ObjectDoesNotExist:
        return redirect('main:ozp')


@login_required(login_url='account:login')
def load_objects(request):
    context = {}
    try:
        ceh = request.GET.get('ceh')
        if ceh == '':
            objects = Object.objects.all()
            context = {'objects': objects, 'ceh': 'все'}
        else:
            department = Department.objects.filter(ceh=ceh, is_prod=True)
            objects = Object.objects.filter(uchastok__in=department)
            context = {'objects': objects, 'ceh': ceh}
    except (ValueError, TypeError):
        pass
    return render(request, 'main/templates/ajax_load_objects.html', context)


@login_required(login_url='account:login')
def load_uchastok(request):
    context = {}
    try:
        ceh = request.GET.get('ceh')
        if ceh == '':
            uchastok = Department.objects.filter(is_prod=True).values_list('uchastok', flat=True)
            context = {'uchastok': uchastok, 'ceh': 'все'}
        else:
            uchastok = Department.objects.filter(ceh=ceh, is_prod=True).values_list('uchastok', flat=True)
            context = {'uchastok': uchastok, 'ceh': ceh}
    except (ValueError, TypeError):
        pass
    return render(request, 'main/templates/ajax_load_uchastok.html', context)


@login_required(login_url='account:login')
def load_object_uchastok(request):
    context = {}
    try:
        uchastok = request.GET.get('uchastok')
        if uchastok == '':
            objects = Object.objects.all()
            context = {'objects': objects, 'uchastok': 'все'}
        else:
            department = Department.objects.filter(uchastok=uchastok, is_prod=True)
            objects = Object.objects.filter(uchastok__in=department)
            context = {'objects': objects, 'uchastok': uchastok}
    except (ValueError, TypeError):
        pass
    return render(request, 'main/templates/ajax_load_objects_uchastok.html', context)


@login_required(login_url='account:login')
@is_staff
def tu_create(request):
    form = Tu_Create_Form(request.POST, request.FILES or None)
    msg = ''
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('main:tu_list')
        else:
            msg = form_errors_text(form)
    context = {'form': form, 'msg': msg}
    return render(request, 'main/templates/tu_creation.html', context)


@login_required(login_url='account:login')
def tu_list(request):
    filter_form = FilterForm()
    object_list = Technicheskie_usloviya.objects.all()
    if request.method == 'GET':
        try:
            data = FilterForm(request.GET or None)
            objects, filtered_by = object_filter(data)
            organization = request.GET.get('organization')
            is_active = request.GET.get('is_active')
            object_list = object_list.filter(object__in=objects)
            if organization != 'all' and organization is not None:
                object_list = object_list.filter(organization=organization)
                filtered_by.append(organization)
            if is_active == '1':
                object_list = object_list.filter(expire_date__gt=datetime.now().date())
                filtered_by.append('Действующим')
            elif is_active == '0':
                object_list = object_list.filter(expire_date__lt=datetime.now().date())
                filtered_by.append('Недействительным')
            filtered_by_str = ' '.join(filtered_by)
            context = {'object_list': object_list, 'filtered_by': filtered_by_str, 'filter_form': filter_form}
            return render(request, 'main/templates/tu_list.html', context)
        except (ValueError, TypeError):
            filtered_by = ''
            context = {'object_list': object_list, 'filtered_by': filtered_by, 'filter_form': filter_form}
            return render(request, 'main/templates/tu_list.html', context)


@login_required(login_url='account:login')
@is_staff
def delete_tu(request, tu_id):
    try:
        obj = Technicheskie_usloviya.objects.get(id=tu_id)
        obj.delete()
        return redirect('main:tu_list')
    except ObjectDoesNotExist:
        return redirect('main:tu_list')


@login_required(login_url='account:login')
def tehdoc_list(request):
    filter_form = FilterForm()
    object_list = Technicheskaya_documentacia.objects.all()
    if request.method == 'GET':
        try:
            data = FilterForm(request.GET or None)
            objects, filtered_by = object_filter(data)
            object_list = object_list.filter(object__in=objects)
            filtered_by_str = ' '.join(filtered_by)
            context = {'object_list': object_list, 'filtered_by': filtered_by_str, 'filter_form': filter_form}
            return render(request, 'main/templates/tehdoc_list.html', context)
        except (ValueError, TypeError):
            filtered_by = ''
            context = {'object_list': object_list, 'filtered_by': filtered_by, 'filter_form': filter_form}
            return render(request, 'main/templates/tehdoc_list.html', context)


@login_required(login_url='account:login')
@is_staff
def delete_tehdoc(request, tehdoc_id):
    try:
        obj = Technicheskaya_documentacia.objects.get(id=tehdoc_id)
        obj.delete()
        return redirect('main:tehdoc_list')
    except ObjectDoesNotExist:
        return redirect('main:tehdoc_list')


@login_required(login_url='account:login')
@is_staff
def tehdoc_create(request):
    form = Tehdoc_Create_Form(request.POST, request.FILES or None)
    msg = ''
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('main:tehdoc_list')
        else:
            msg = form_errors_text(form)
    context = {'form': form, 'msg': msg}
    return render(request, 'main/templates/tehdoc_creation.html', context)


@login_required(login_url='account:login')
def scheme_list(request):
    filter_form = FilterForm()
    object_list = Scheme.objects.all()
    if request.method == 'GET':
        try:
            data = FilterForm(request.GET or None)
            objects, filtered_by = object_filter(data)
            object_list = object_list.filter(object__in=objects)
            filtered_by_str = ' '.join(filtered_by)
            context = {'object_list': object_list, 'filtered_by': filtered_by_str, 'filter_form': filter_form}
            return render(request, 'main/templates/scheme_list.html', context)
        except (ValueError, TypeError):
            filtered_by = ''
            context = {'object_list': object_list, 'filtered_by': filtered_by, 'filter_form': filter_form}
            return render(request, 'main/templates/scheme_list.html', context)


@login_required(login_url='account:login')
@is_staff
def scheme_create(request):
    form = Scheme_Create_Form(request.POST, request.FILES or None)
    msg = ''
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('main:scheme_list')
        else:
            msg = form_errors_text(form)
    context = {'form': form, 'msg': msg}
    return render(request, 'main/templates/scheme_creation.html', context)


@login_required(login_url='account:login')
@is_staff
def scheme_delete(request, scheme_id):
    try:
        obj = Scheme.objects.get(id=scheme_id)
        obj.delete()
        return redirect('main:scheme_list')
    except ObjectDoesNotExist:
        return redirect('main:scheme_list')


@login_required(login_url='account:login')
@is_staff
def change_object_name(request, obj_id: int):
    try:
        obj = Object.objects.get(id=obj_id)
        form = Object_Name_Form(request.POST)
        if request.method == 'POST':
            if form.is_valid():
                name = form.cleaned_data['object_name']
                obj.object_name = name
                obj.save()
                return redirect('main:object_detail', name=obj_id)
            else:
                msg = form_errors_text(form)
                return redirect('main:object_detail', name=obj_id, msg=msg)
    except ObjectDoesNotExist:
        msg = 'Неизвестная ошибка'
        return redirect('main:object_detail', name=obj_id, msg=msg)


@login_required(login_url='account:login')
@is_staff
def change_object_uchastok(request, obj_id: int):
    try:
        obj = Object.objects.get(id=obj_id)
        form = Object_Uchastok_Form(request.POST)
        if request.method == 'POST':
            if form.is_valid():
                uchastok = form.cleaned_data['uchastok']
                obj.uchastok = uchastok
                obj.save()
                return redirect('main:object_detail', name=obj_id)
            else:
                msg = form_errors_text(form)
                return redirect('main:object_detail', name=obj_id, msg=msg)
    except ObjectDoesNotExist:
        msg = 'Неизвестная ошибка'
        return redirect('main:object_detail', name=obj_id, msg=msg)


@login_required(login_url='account:login')
@is_staff
def change_coordinats(request, obj_id: int):
    try:
        obj = Object.objects.get(id=obj_id)
        form = Coordinat_Form(request.POST)
        if request.method == 'POST':
            if form.is_valid():
                obj.position.latitude_degrees = form.cleaned_data['latitude_degrees']
                obj.position.latitude_minutes = form.cleaned_data['latitude_minutes']
                obj.position.latitude_seconds = form.cleaned_data['latitude_seconds']
                obj.position.longitude_degrees = form.cleaned_data['longitude_degrees']
                obj.position.longitude_minutes = form.cleaned_data['longitude_minutes']
                obj.position.longitude_seconds = form.cleaned_data['longitude_seconds']
                obj.position.save()
                return redirect('main:object_detail', name=obj_id)
            else:
                msg = form_errors_text(form)
                return redirect('main:object_detail', name=obj_id, msg=msg)
    except ObjectDoesNotExist:
        msg = 'Неизвестная ошибка'
        return redirect('main:object_detail', name=obj_id, msg=msg)


@login_required(login_url='account:login')
@is_staff
def change_address(request, obj_id: int):
    try:
        obj = Object.objects.get(id=obj_id)
        form = Object_Address_Form(request.POST)
        if request.method == 'POST':
            if form.is_valid():
                obj.position.address = form.cleaned_data['address']
                obj.position.save()
                return redirect('main:object_detail', name=obj_id)
            else:
                msg = form_errors_text(form)
                return redirect('main:object_detail', name=obj_id, msg=msg)
    except ObjectDoesNotExist:
        msg = 'Неизвестная ошибка'
        return redirect('main:object_detail', name=obj_id, msg=msg)


@login_required(login_url='account:login')
@is_staff
def change_district(request, obj_id: int):
    try:
        obj = Object.objects.get(id=obj_id)
        form = Object_District_Form(request.POST)
        if request.method == 'POST':
            if form.is_valid():
                obj.position.district = form.cleaned_data['district']
                obj.position.save()
                return redirect('main:object_detail', name=obj_id)
            else:
                msg = form_errors_text(form)
                return redirect('main:object_detail', name=obj_id, msg=msg)
    except ObjectDoesNotExist:
        msg = 'Неизвестная ошибка'
        return redirect('main:object_detail', name=obj_id, msg=msg)


@login_required(login_url='account:login')
@is_staff
def add_object_foto(request, obj_id: int):
    try:
        obj = Object.objects.get(id=obj_id)
        form = Object_Foto_Form(request.POST, request.FILES)
        if request.method == 'POST':
            if form.is_valid():
                year = form.cleaned_data['year']
                foto = request.FILES['foto']
                foto_object = Object_Foto.objects.create(year=year, foto=foto, object=obj)
                foto_object.save()
                return redirect('main:object_detail', name=obj_id)
            else:
                msg = form_errors_text(form)
                return redirect('main:object_detail', name=obj_id, msg=msg)
    except ObjectDoesNotExist:
        msg = 'Неизвестная ошибка'
        return redirect('main:object_detail', name=obj_id, msg=msg)


@login_required(login_url='account:login')
@is_staff
def foto_object_delete(request, foto_id: int):
    try:
        foto_object = Object_Foto.objects.get(id=foto_id)
        object_id = foto_object.object.id
        foto_object.delete()
        return redirect('main:object_detail', name=object_id)
    except ObjectDoesNotExist:
        msg = 'Неизвестная ошибка'
        return redirect('main:objects', msg=msg)


@login_required(login_url='account:login')
@is_staff
def delete_object(request, obj_id: int):
    try:
        obj = Object.objects.get(id=obj_id)
        name = str(obj.object_name)
        obj.delete()
        return redirect('main:objects', msg=f'Объект {name} удален')
    except ObjectDoesNotExist:
        msg = 'Неизвестная ошибка'
        return redirect('main:objects', msg=msg)


@login_required(login_url='account:login')
def object_foto_list(request, msg=''):
    foto_list = Object_Foto.objects.all()
    foto_form = List_Foto_Form()
    filter_form = FilterForm()
    if request.method == 'GET':
        try:
            data = FilterForm(request.GET or None)
            objects, filtered_by = object_filter(data)
            filtered_by_str = ' '.join(filtered_by)
            foto_list = foto_list.filter(object__in=objects)
            context = {'foto_list': foto_list, 'filtered_by': filtered_by_str, 'filter_form': filter_form, 'msg': msg,
                       'foto_form': foto_form}
            return render(request, 'main/templates/foto_object_list.html', context)
        except (ValueError, TypeError):
            filtered_by = ''
            context = {'foto_list': foto_list, 'filtered_by': filtered_by, 'filter_form': filter_form, 'msg': msg,
                       'foto_form': foto_form}
            return render(request, 'main/templates/foto_object_list.html', context)


@login_required(login_url='account:login')
@is_staff
def foto_add_list(request):
    foto_form = List_Foto_Form(request.POST, request.FILES or None)
    if request.method == 'POST':
        if foto_form.is_valid():
            foto_form.save()
            return redirect('main:object_foto_list')
        else:
            msg = form_errors_text(foto_form)
            return redirect('main:object_foto_list', msg=msg)


@login_required(login_url='account:login')
@is_staff
def foto_list_delete(request, foto_id):
    try:
        obj = Object_Foto.objects.get(id=foto_id)
        name = str(obj.object.object_name)
        obj.delete()
        return redirect('main:object_foto_list', msg=f'Фотография объекта {name} удалена')
    except ObjectDoesNotExist:
        msg = 'Неизвестная ошибка'
        return redirect('main:object_foto_list', msg=msg)
