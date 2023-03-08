from pathlib import Path
from datetime import date, timedelta, datetime

from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from openpyxl.styles import Border, Side, Alignment, Font
from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl import Workbook

from main.scripts import form_errors_text, is_staff
from main.views import department_filter
from main.models import Object, Department
from rich.forms import ResForm, RichForm, RegistrationForm, ResNameForm, ResClassForm, RegForm, TypeForm, \
    ResProtokolForm
from rich.models import Rich, Res, Registration, Type, ResProtokol


@login_required(login_url='account:login')
def rich_app_index(request):
    return render(request, "rich/templates/rich_app_index.html")


@login_required(login_url='account:login')
def rich_index(request):
    # получение списков данных для меню выбора -------------------
    dep = Department.objects.all()
    ceh_list = []
    uchastok_list = []
    for i in dep:
        if i.ceh in ceh_list:
            pass
        else:
            ceh_list.append(i.ceh)
        if i.uchastok in uchastok_list:
            pass
        else:
            uchastok_list.append(i.uchastok)
    type = Type.objects.all()
    # --------------------------------------------------------------
    # получение введенных в фильр данных -------------------------------------------
    if request.GET:
        ceh = request.GET['ceh']
        uchastok = request.GET['uchastok']
        object = request.GET['object']
        is_active = request.GET['is_active']
        type_res = request.GET['type_res']
        equip_name = request.GET['equip_name']
        frequency_range = request.GET['frequency_range']
        # фильтрация по поздразделениям-----------------------------------------------
        filtered_by = []
        # пустой список для вывода критериев фильтрации
        if ceh == 'all' and uchastok == 'all':
            filtered_res_objects = Res.objects.all()
            object_list = Object.objects.all().select_related('related_object', 'related_rich', 'related_registration',
                                                              'type').prefetch_related('protokol')
        else:
            filtered_departments = department_filter(ceh, uchastok)[0]
            filtered_by.extend(department_filter(ceh, uchastok)[1])
            filtered_objects = Object.objects.filter(uchastok__in=filtered_departments)
            object_list = filtered_objects
            filtered_res_objects = Res.objects.filter(related_object__in=filtered_objects)
        # фильтрация по объектам сети-----------------------------------------------
        if object != 'all':
            filtered_object = Object.objects.filter(object_name=object)
            filtered_res_objects = Res.objects.filter(related_object__in=filtered_object).\
                select_related('related_object', 'related_rich','related_registration', 'type').\
                prefetch_related('protokol')
            filtered_by.append(str(object))
        else:
            pass
        # фильтрация по активным -----------------------------------------------------
        if is_active == '1':
            filtered_by.append('Действующим')
            filtered_res_objects = filtered_res_objects.filter(is_active=True)
        elif is_active == '0':
            filtered_by.append('Проектируемым')
            filtered_res_objects = filtered_res_objects.filter(is_active=False)
        else:
            filtered_by_is_active = ''
        # -----------------------------------------------------------------------------
        # фильтрация по типу -----------------------------------------------------
        if type_res == 'all':
            pass
        else:
            filtered_by.append(str(type_res))
            filtered_by_type_set = Type.objects.filter(class_name=type_res)
            filtered_res_objects = filtered_res_objects.filter(type__in=filtered_by_type_set)
        # -----------------------------------------------------------------------------
        # фильтрация по наименованию оборудования -------------------------------------
        if equip_name == 'all':
            pass
        else:
            filtered_by.append(str(equip_name))
            filtered_by_equipment_set = Type.objects.filter(name=equip_name)
            filtered_res_objects = filtered_res_objects.filter(type__in=filtered_by_equipment_set)
        # -----------------------------------------------------------------------------
        # фильтрация по диапазону -----------------------------------------------------
        if frequency_range == 'all':
            pass
        else:
            filtered_by.append('Диапазону ' + str(frequency_range))
            filtered_by_frequency_set = Type.objects.filter(frequency_str=frequency_range)
            filtered_res_objects = filtered_res_objects.filter(type__in=filtered_by_frequency_set)
        # -----------------------------------------------------------------------------
        filtered_by_str = ' '.join(filtered_by)
        context = {'objects': object_list, 'ceh': ceh_list, 'uchastok': uchastok_list, 'filtered_by': filtered_by_str,
                   'object': filtered_res_objects, 'type_list': type}
        return render(request, "rich/templates/rich_index.html", context)
    else:
        filtered_res_objects = Res.objects.all().select_related('related_object', 'related_rich','related_registration',
                                                                'type').prefetch_related('protokol')
        object_list = Object.objects.all()
        filtered_by = ''
        context = {'objects': object_list, 'ceh': ceh_list, 'uchastok': uchastok_list, 'filtered_by': filtered_by,
                   'object': filtered_res_objects, 'type_list': type}
        return render(request, "rich/templates/rich_index.html", context)


@login_required(login_url='account:login')
@is_staff
def res_creation(request):
    form = ResForm(request.POST or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return rich_index(request)
        else:
            msg = form_errors_text(form)
            return render(request, "rich/templates/res_creation.html", {'form': form, 'msg': msg})
    return render(request, "rich/templates/res_creation.html", {'form': form})


@login_required(login_url='account:login')
def res_detail(request, id, msg=''):
    res_object = Res.objects.select_related('related_object', 'related_rich', 'related_registration').get(id=id)
    related_object = Object.objects.all()
    related_rich = Rich.objects.all()
    related_registration = Registration.objects.all()
    rich_form = RichForm()
    registration_form = RegistrationForm()
    res_form = ResForm()
    type = Type.objects.all()
    type_form = TypeForm()
    protokol_form = ResProtokolForm()
    context = {'object': res_object, 'related_object': related_object, 'related_rich': related_rich,
               'rich_form': rich_form, 'related_registration': related_registration, 'type': type,
               'registration_form': registration_form, 'msg': msg, 'res_form': res_form, 'type_form': type_form,
               'protokol_form': protokol_form}
    return render(request, "rich/templates/res_details.html", context)


@login_required(login_url='account:login')
@is_staff
def change_name(request, id):
    res_object = Res.objects.get(id=id)
    f = ResNameForm(request.POST or None)
    if request.method == 'POST':
        if f.is_valid():
            res_object.name = f.cleaned_data['name']
            res_object.save()
        else:
            msg = form_errors_text(f)
            return res_detail(request, id, msg=msg)
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def change_object(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        obj_name = request.POST['select_object']
        obj = Object.objects.get(object_name=obj_name)
        res_object.related_object = obj
        res_object.save()
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def change_rich(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        rich_name = request.POST['select_rich']
        obj = Rich.objects.get(name=rich_name)
        if obj.end_date < date.today():
            return res_detail(request, id, msg='Вы не можете присвоить РИЧ с истекшим сроком действия!')
        else:
            res_object.related_rich = obj
            res_object.save()
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def add_rich(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        rich_form = RichForm(request.POST, request.FILES)
        file = request.FILES['doc']
        name = file.name.replace(Path(file.name).suffix, '')
        name = name.replace('_', ' ')
        end_date = request.POST['end_date']
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        if rich_form.is_valid() and (Rich.objects.filter(name=name).exists() is False) and (end_date > date.today()):
            a = rich_form.save()
            res_object.related_rich = a
            res_object.save()
            return redirect('rich:res_detail', id=id)
        elif Rich.objects.filter(name=name).exists():
            return res_detail(request, id, msg='Разрешение с таким именем уже существует!')
        elif end_date < date.today():
            return res_detail(request, id, msg='Вы не можете добавить РИЧ с истекшим сроком действия!')
        else:
            msg = form_errors_text(rich_form)
            return res_detail(request, id, msg=msg)
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def change_registration(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        reg = request.POST['select_registration']
        obj = Registration.objects.get(name=reg)
        if obj.related_res_reg is not None:
            return res_detail(request, id, msg='На указанный документ уже зарегистрирован РЭС!')
        elif obj.end_date < date.today():
            return res_detail(request, id, msg='Вы не можете привязать документ с истекшим сроком действия!')
        else:
            res_object.related_registration = obj
            res_object.save()
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def add_registration(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        registration_form = RegistrationForm(request.POST, request.FILES)
        file = request.FILES['doc']
        name = file.name.replace(Path(file.name).suffix, '')
        name = name.replace('_', ' ')
        end_date = request.POST['end_date']
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        if registration_form.is_valid() and (Registration.objects.filter(name=name).exists() is False) and \
                (res_object.related_rich is not None) and (end_date > date.today()):
            a = registration_form.save()
            res_object.related_registration = a
            res_object.save()
            return redirect('rich:res_detail', id=id)
        elif Registration.objects.filter(name=name).exists():
            return res_detail(request, id, msg='Регистрационный документ с таким именем уже существует!')
        elif res_object.related_rich is None:
            return res_detail(request, id, msg='Сначала добавьте РИЧ!')
        elif end_date < date.today():
            return res_detail(request, id, msg='Вы не можете добавить документ с истекшим сроком действия!')
        else:
            msg = form_errors_text(registration_form)
            return res_detail(request, id, msg=msg)
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def change_radio_class(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        f = ResClassForm(request.POST)
        if f.is_valid():
            res_object.radio_class = f.cleaned_data['radio_class']
            res_object.save()
        else:
            return res_detail(request, id, msg='Вы ввели некорректные данные!')
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def antenna_height_change(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        height = request.POST['antenna_heigh']
        if height == '':
            res_object.antenna_height = None
        else:
            res_object.antenna_height = height
        res_object.save()
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def antenna_azimuth_change(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        azimuth = request.POST['antenna_azimuth']
        if azimuth == '':
            res_object.azimuth_value = None
        else:
            res_object.azimuth_value = azimuth
        res_object.save()
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def antenna_gain_change(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        gain = request.POST['antenna_gain']
        if gain == '':
            res_object.antenna_gain = None
        else:
            res_object.antenna_gain = gain
        res_object.save()
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
@is_staff
def antenna_polarization_change(request, id):
    res_object = Res.objects.get(id=id)
    if request.method == 'POST':
        polarization = request.POST['polarization']
        if polarization == '':
            res_object.polarization = None
        else:
            res_object.polarization = polarization
        res_object.save()
    return redirect('rich:res_detail', id=id)


@login_required(login_url='account:login')
def rich_list(request, msg=''):
    # получение списков данных для меню выбора -------------------
    dep = Department.objects.all()
    ceh_list = []
    uchastok_list = []
    for i in dep:
        if i.ceh in ceh_list:
            pass
        else:
            ceh_list.append(i.ceh)
        if i.uchastok in uchastok_list:
            pass
        else:
            uchastok_list.append(i.uchastok)
    # --------------------------------------------------------------
    # получение введенных в фильр данных -------------------------------------------
    if request.GET:
        ceh = request.GET['ceh']
        uchastok = request.GET['uchastok']
        object = request.GET['object']
        is_active = request.GET['is_active']
        srok_this_year = request.GET['srok_this_year']
        # фильтрация по поздразделениям-----------------------------------------------
        if ceh == 'all' and uchastok == 'all':
            filtered_rich_objects = Rich.objects.all()
            filtered_by_department = ''
            object_list = Object.objects.all()
        else:
            filtered_departments = department_filter(ceh, uchastok)[0]
            filtered_by_department = " ".join(department_filter(ceh, uchastok)[1])
            filtered_objects = Object.objects.filter(uchastok__in=filtered_departments)
            object_list = filtered_objects
            filtered_res = Res.objects.filter(related_object__in=filtered_objects)
            filtered_rich_objects = Rich.objects.filter(related_res__in=filtered_res).distinct()
        # фильтрация по объектам сети-----------------------------------------------
        if object != 'all':
            filtered_object = Object.objects.filter(object_name=object)
            filtered_res = Res.objects.filter(related_object__in=filtered_object)
            filtered_rich_objects = Rich.objects.filter(related_res__in=filtered_res).distinct()
            filtered_by_object = object
        else:
            filtered_by_object = ''
        # фильтрация по активным -----------------------------------------------------
        if is_active == '1':
            filtered_by_is_active = 'Действующим'
            filtered_rich_objects = filtered_rich_objects.filter(is_active=True)
        elif is_active == '0':
            filtered_by_is_active = 'Недействительным'
            filtered_rich_objects = filtered_rich_objects.filter(is_active=False)
        else:
            filtered_by_is_active = ''
        # -----------------------------------------------------------------------------
        # фильтрация сроку в течении года ----------------------------------------
        if srok_this_year == '1':
            filtered_by_srok_this_year = 'Сроку окончания действия менее года'
            filtered_rich_objects = filtered_rich_objects.filter(end_date__lt=(date.today() + timedelta(days=365)))
        else:
            filtered_by_srok_this_year = ''
        # -------------------------------------------------------------------
        filtered_by = ' '.join([filtered_by_department, filtered_by_object, filtered_by_is_active,
                                filtered_by_srok_this_year])
        if filtered_by != '    ':
            pass
        else:
            filtered_by = ''
        context = {'objects': object_list, 'ceh': ceh_list, 'uchastok': uchastok_list, 'filtered_by': filtered_by,
                   'object': filtered_rich_objects}
        return render(request, "rich/templates/rich_list.html", context)
    else:
        object = Rich.objects.all().prefetch_related()
        filtered_by = str('')
        objects_list = Object.objects.all()
        context = {'objects': objects_list, 'ceh': ceh_list, 'uchastok': uchastok_list, 'filtered_by': filtered_by,
                   'object': object, 'msg': msg}
        return render(request, "rich/templates/rich_list.html", context)


@login_required(login_url='account:login')
@is_staff
def rich_create(request):
    if request.method == 'POST':
        rich_form = RichForm(request.POST, request.FILES)
        if rich_form.is_valid():
            d = rich_form.save(commit=False)
            file = request.FILES['doc']
            name = file.name.replace(Path(file.name).suffix, '')
            name = name.replace('_', ' ')
            end_date = request.POST['end_date']
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            if Rich.objects.filter(name=name).exists():
                msg = 'РИЧ с таким именем уже существует!'
            elif end_date < date.today():
                msg = 'Вы не можете создать РИЧ с истекшим сроком действия!'
            else:
                d.name = name
                d.save()
                return redirect('rich:rich_list')
        else:
            msg = 'Вы ввели некорректные данные!'
        form = RichForm()
        context = {'msg': msg, 'form': form}
        return render(request, "rich/templates/rich_creation.html", context)
    else:
        form = RichForm()
        context = {'form': form}
        return render(request, "rich/templates/rich_creation.html", context)


@login_required(login_url='account:login')
@is_staff
def delete_rich(request, id):
    rich_object = Rich.objects.get(id=id)
    rich_object.delete()
    return redirect('rich:rich_list')


@login_required(login_url='account:login')
def export_xls_rich(request):
    if request.method == 'GET':
        try:
            ceh = request.GET['ceh']
            uchastok = request.GET['uchastok']
            object = request.GET['object']
            is_active = request.GET['is_active']
            srok_this_year = request.GET['srok_this_year']
            # фильтрация по поздразделениям-----------------------------------------------
            if ceh == 'all' and uchastok == 'all':
                filtered_rich_objects = Rich.objects.all().prefetch_related('related_res', )
            else:
                filtered_departments = department_filter(ceh, uchastok)[0]
                filtered_objects = Object.objects.filter(uchastok__in=filtered_departments)
                filtered_res = Res.objects.filter(related_object__in=filtered_objects)
                filtered_rich_objects = Rich.objects.filter(related_res__in=filtered_res).distinct()
            # фильтрация по объектам сети-----------------------------------------------
            if object != 'all':
                filtered_object = Object.objects.filter(object_name=object)
                filtered_res = Res.objects.filter(related_object__in=filtered_object)
                filtered_rich_objects = Rich.objects.filter(related_res__in=filtered_res).distinct()
            else:
                pass
            # фильтрация по активным -----------------------------------------------------
            if is_active == '1':
                filtered_rich_objects = filtered_rich_objects.filter(is_active=True)
            elif is_active == '0':
                filtered_rich_objects = filtered_rich_objects.filter(is_active=False)
            else:
                pass
            # -----------------------------------------------------------------------------
            # фильтрация сроку в течении года ----------------------------------------
            if srok_this_year == '1':
                filtered_rich_objects = filtered_rich_objects.filter(end_date__lt=(date.today() + timedelta(days=365)))
            else:
                pass
        except KeyError:
            filtered_rich_objects = Rich.objects.all()
        # выгрузка в xls -----------------------------------------------------
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="rich.xlsx"'
        wb = Workbook()
        rows = filtered_rich_objects
        ws = wb.active
        row_num = 1
        columns = ['РИЧ', 'Дата начала действия РИЧ', 'Дата окончания РИЧ', 'Срок действия РИЧ', 'Перечень РЭС']
        font = Font(name='TimesNewRoman', sz=11, bold=True)
        font2 = Font(name='TimesNewRoman', sz=11, bold=False)
        border = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'))
        alignment = Alignment(horizontal='center',
                              vertical='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
        for col_num in range(len(columns)):
            ws.cell(row_num, col_num + 1, columns[col_num])
            ws.cell(row_num, col_num + 1).font = font
            ws.cell(row_num, col_num + 1).border = border
            ws.cell(row_num, col_num + 1).alignment = alignment
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        for row in range(rows.count()):
            if rows[row].is_active:
                x = rows[row].days_left()
            else:
                x = 'РИЧ недействителен'
            y = []
            z = ''
            for i in rows[row].related_res.all():
                y.append(i.name)
            z = " ".join(y)
            atr_list = [rows[row].name, rows[row].start_date, rows[row].end_date, x, z]
            for col_num in range(len(atr_list)):
                ws.cell(row + 2, col_num + 1, atr_list[col_num])
                ws.cell(row + 2, col_num + 1).font = font2
                ws.cell(row + 2, col_num + 1).border = border
                ws.cell(row + 2, col_num + 1).alignment = alignment
        wb.save(response)
        return response


@login_required(login_url='account:login')
def reg_list(request):
    # получение списков данных для меню выбора -------------------
    dep = Department.objects.all()
    ceh_list = []
    uchastok_list = []
    for i in dep:
        if i.ceh in ceh_list:
            pass
        else:
            ceh_list.append(i.ceh)
        if i.uchastok in uchastok_list:
            pass
        else:
            uchastok_list.append(i.uchastok)
    # --------------------------------------------------------------
    # получение введенных в фильр данных -------------------------------------------
    if request.GET:
        ceh = request.GET['ceh']
        uchastok = request.GET['uchastok']
        object = request.GET['object']
        is_active = request.GET['is_active']
        srok_this_month = request.GET['srok_this_month']
        # фильтрация по поздразделениям-----------------------------------------------
        if ceh == 'all' and uchastok == 'all':
            filtered_reg_objects = Registration.objects.all()
            filtered_by_department = ''
            object_list = Object.objects.all()
        else:
            filtered_departments = department_filter(ceh, uchastok)[0]
            filtered_by_department = " ".join(department_filter(ceh, uchastok)[1])
            filtered_objects = Object.objects.filter(uchastok__in=filtered_departments)
            object_list = filtered_objects
            filtered_res = Res.objects.filter(related_object__in=filtered_objects)
            filtered_reg_objects = Registration.objects.filter(related_res_reg__in=filtered_res).distinct()
        # фильтрация по объектам сети-----------------------------------------------
        if object != 'all':
            filtered_object = Object.objects.filter(object_name=object)
            filtered_res = Res.objects.filter(related_object__in=filtered_object)
            filtered_reg_objects = Registration.objects.filter(related_res_reg__in=filtered_res).distinct()
            filtered_by_object = object
        else:
            filtered_by_object = ''
        # фильтрация по активным -----------------------------------------------------
        if is_active == '1':
            filtered_by_is_active = 'Действующим'
            filtered_reg_objects = filtered_reg_objects.filter(is_active=True)
        elif is_active == '0':
            filtered_by_is_active = 'Недействительным'
            filtered_reg_objects = filtered_reg_objects.filter(is_active=False)
        else:
            filtered_by_is_active = ''
        # -----------------------------------------------------------------------------
        # фильтрация сроку в течении года ----------------------------------------
        if srok_this_month == '1':
            filtered_by_srok_this_month = 'Сроку окончания действия менее месяца'
            filtered_reg_objects = filtered_reg_objects.filter(end_date__lt=(date.today() + timedelta(days=30)))
        else:
            filtered_by_srok_this_month = ''
        # -------------------------------------------------------------------
        filtered_by = ' '.join([filtered_by_department, filtered_by_object, filtered_by_is_active,
                                filtered_by_srok_this_month])
        if filtered_by != '    ':
            pass
        else:
            filtered_by = ''
        context = {'objects': object_list, 'ceh': ceh_list, 'uchastok': uchastok_list, 'filtered_by': filtered_by,
                   'object': filtered_reg_objects}
        return render(request, "rich/templates/reg_list.html", context)
    reg_object = Registration.objects.prefetch_related('related_res_reg').all()
    object_list = Object.objects.all()
    context = {'object': reg_object, 'ceh': ceh_list, 'uchastok': uchastok_list, 'objects': object_list}
    return render(request, "rich/templates/reg_list.html", context)


@login_required(login_url='account:login')
def export_xls_reg(request):
    if request.method == 'GET':
        try:
            ceh = request.GET['ceh']
            uchastok = request.GET['uchastok']
            object = request.GET['object']
            is_active = request.GET['is_active']
            srok_this_month = request.GET['srok_this_month']
            # фильтрация по поздразделениям-----------------------------------------------
            if ceh == 'all' and uchastok == 'all':
                filtered_reg_objects = Registration.objects.all().prefetch_related('related_res')
            else:
                filtered_departments = department_filter(ceh, uchastok)[0]
                filtered_objects = Object.objects.filter(uchastok__in=filtered_departments)
                filtered_res = Res.objects.filter(related_object__in=filtered_objects)
                filtered_reg_objects = Registration.objects.filter(related_res_reg__in=filtered_res).distinct()
            # фильтрация по объектам сети-----------------------------------------------
            if object != 'all':
                filtered_object = Object.objects.filter(object_name=object)
                filtered_res = Res.objects.filter(related_object__in=filtered_object)
                filtered_reg_objects = Registration.objects.filter(related_res_reg__in=filtered_res).distinct()
            else:
                pass
            # фильтрация по активным -----------------------------------------------------
            if is_active == '1':
                filtered_reg_objects = filtered_reg_objects.filter(is_active=True)
            elif is_active == '0':
                filtered_reg_objects = filtered_reg_objects.filter(is_active=False)
            else:
                pass
            # -----------------------------------------------------------------------------
            # фильтрация сроку в течении года ----------------------------------------
            if srok_this_month == '1':
                filtered_reg_objects = filtered_reg_objects.filter(end_date__lt=(date.today() + timedelta(days=365)))
            else:
                pass
        except KeyError:
            filtered_reg_objects = Registration.objects.all()
        # выгрузка в xls -----------------------------------------------------
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="registration.xlsx"'
        wb = Workbook()
        rows = filtered_reg_objects
        ws = wb.active
        row_num = 1
        columns = ['Регистрация', 'Дата начала действия', 'Дата окончания', 'Срок действия', 'РЭС', 'РИЧ']
        font = Font(name='TimesNewRoman', sz=11, bold=True)
        font2 = Font(name='TimesNewRoman', sz=11, bold=False)
        border = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'))
        alignment = Alignment(horizontal='center',
                              vertical='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
        for col_num in range(len(columns)):
            ws.cell(row_num, col_num + 1, columns[col_num])
            ws.cell(row_num, col_num + 1).font = font
            ws.cell(row_num, col_num + 1).border = border
            ws.cell(row_num, col_num + 1).alignment = alignment
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        for row in range(rows.count()):
            atr_list = [rows[row].name, rows[row].start_date, rows[row].end_date, rows[row].days_left(),
                        rows[row].related_res_reg.name, rows[row].related_res_reg.related_rich.name]
            for col_num in range(len(atr_list)):
                ws.cell(row + 2, col_num + 1, value=atr_list[col_num])
                ws.cell(row + 2, col_num + 1).font = font2
                ws.cell(row + 2, col_num + 1).border = border
                ws.cell(row + 2, col_num + 1).alignment = alignment
        wb.save(response)
        return response


@login_required(login_url='account:login')
@is_staff
def delete_reg(request, key):
    reg_object = Registration.objects.get(id=key)
    reg_object.delete()
    return redirect('rich:reg_list')


@login_required(login_url='account:login')
@is_staff
def reg_create(request):
    if request.method == 'POST':
        reg_form = RegForm(request.POST, request.FILES)
        if reg_form.is_valid():
            d = reg_form.save(commit=False)
            file = request.FILES['doc']
            related_res = request.POST['related_res']
            res_obj = Res.objects.get(id=related_res)
            end_date = request.POST['end_date']
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
            if res_obj.related_registration is not None:
                msg = 'Указанный РЭС уже зарегистирован!'
            elif res_obj.related_rich is None:
                msg = 'Для указанного РЭС не привязан РИЧ!'
            elif end_date < date.today():
                msg = 'Вы не можете создать документ с истекшим сроком действия!'
            else:
                name = file.name.replace(Path(file.name).suffix, '')
                name = name.replace('_', ' ')
                if Registration.objects.filter(name=name).exists():
                    msg = 'Регистрация с таким именем уже существует!'
                else:
                    d.name = name
                    d.save()
                    res_obj.related_registration = d
                    res_obj.save()
                    return redirect('rich:reg_list')
            form = RegForm()
            context = {'msg': msg, 'form': form}
            return render(request, "rich/templates/reg_creation.html", context)
        else:
            msg = 'Вы ввели некорректные данные!'
        form = RegForm()
        context = {'msg': msg, 'form': form}
        return render(request, "rich/templates/reg_creation.html", context)
    else:
        form = RegForm()
        context = {'form': form}
        return render(request, "rich/templates/reg_creation.html", context)


@login_required(login_url='account:login')
@is_staff
def load_res(request):
    object_id = request.GET.get('related_object')
    res_list = Res.objects.filter(related_object=object_id, related_registration=None)
    return render(request, 'rich/templates/res_list_options.html', {'res_list': res_list})


@login_required(login_url='account:login')
def res_export_xls(request):
    if request.method == 'GET':
        try:
            ceh = request.GET['ceh']
            uchastok = request.GET['uchastok']
            object = request.GET['object']
            is_active = request.GET['is_active']
            type_res = request.GET['type_res']
            equip_name = request.GET['equip_name']
            frequency_range = request.GET['frequency_range']
            # фильтрация по поздразделениям-----------------------------------------------
            if ceh == 'all' and uchastok == 'all':
                filtered_res_objects = Res.objects.all().select_related('related_rich', 'related_registration', 'type')
            else:
                filtered_departments = department_filter(ceh, uchastok)[0]
                filtered_objects = Object.objects.filter(uchastok__in=filtered_departments)
                filtered_res_objects = Res.objects.filter(related_object__in=filtered_objects)
            # фильтрация по объектам сети-----------------------------------------------
            if object != 'all':
                filtered_object = Object.objects.filter(object_name=object)
                filtered_res_objects = Res.objects.filter(related_object__in=filtered_object)
            else:
                pass
            # фильтрация по активным -----------------------------------------------------
            if is_active == '1':
                filtered_res_objects = filtered_res_objects.filter(is_active=True)
            elif is_active == '0':
                filtered_res_objects = filtered_res_objects.filter(is_active=False)
            else:
                pass
            # -----------------------------------------------------------------------------
            # фильтрация по типу -----------------------------------------------------
            if type_res == 'all':
                pass
            else:
                filtered_by_type_set = Type.objects.filter(class_name=type_res)
                filtered_res_objects = filtered_res_objects.filter(type__in=filtered_by_type_set)
            # -----------------------------------------------------------------------------
            # фильтрация по наименованию оборудования -------------------------------------
            if equip_name == 'all':
                pass
            else:
                filtered_by_equipment_set = Type.objects.filter(name=equip_name)
                filtered_res_objects = filtered_res_objects.filter(type__in=filtered_by_equipment_set)
            # -----------------------------------------------------------------------------
            # фильтрация по диапазону -----------------------------------------------------
            if frequency_range == 'all':
                pass
            else:
                filtered_by_frequency_set = Type.objects.filter(frequency_str=frequency_range)
                filtered_res_objects = filtered_res_objects.filter(type__in=filtered_by_frequency_set)
            # -----------------------------------------------------------------------------
        except KeyError:
            filtered_res_objects = Res.objects.all()
        # выгрузка в xls -----------------------------------------------------
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="res_list.xlsx"'
        wb = Workbook()
        rows = filtered_res_objects
        print(rows)
        ws = wb.active
        row_num = 1
        columns = ['РЭС', 'Тип РЭС', 'Оборудование', 'Объект', 'РИЧ', 'Срок действия РИЧ', 'Регистрация',
                   'Срок действия регистрации', 'В работе']
        font = Font(name='TimesNewRoman', sz=11, bold=True)
        font2 = Font(name='TimesNewRoman', sz=11, bold=False)
        border = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'))
        alignment = Alignment(horizontal='center',
                              vertical='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
        for col_num in range(len(columns)):
            ws.cell(row_num, col_num + 1, columns[col_num])
            ws.cell(row_num, col_num + 1).font = font
            ws.cell(row_num, col_num + 1).border = border
            ws.cell(row_num, col_num + 1).alignment = alignment
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 30
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 30
        ws.column_dimensions['H'].width = 30
        ws.column_dimensions['I'].width = 30
        for row in range(rows.count()):
            atr_list = [rows[row].name]
            try:
                atr_list.append(rows[row].type.class_name)
            except AttributeError:
                atr_list.append('-')
            try:
                atr_list.append(rows[row].type.manufacter)
            except AttributeError:
                atr_list.append('-')
            atr_list.append(rows[row].related_object.object_name)
            try:
                atr_list.append(rows[row].related_rich.name)
            except AttributeError:
                atr_list.append('-')
            try:
                atr_list.append(rows[row].related_rich.days_left())
            except AttributeError:
                atr_list.append('-')
            try:
                atr_list.append(rows[row].related_registration.name)
            except AttributeError:
                atr_list.append('-')
            try:
                atr_list.append(rows[row].related_registration.days_left())
            except AttributeError:
                atr_list.append('-')
            atr_list.append(rows[row].active_status())
            for col_num in range(len(atr_list)):
                ws.cell(row + 2, col_num + 1, value=atr_list[col_num])
                ws.cell(row + 2, col_num + 1).font = font2
                ws.cell(row + 2, col_num + 1).border = border
                ws.cell(row + 2, col_num + 1).alignment = alignment
        wb.save(response)
        return response


@login_required(login_url='account:login')
@is_staff
def delete_res(request, key):
    res_object = Res.objects.get(id=key)
    res_object.delete()
    return redirect('rich:rich_index')


@login_required(login_url='account:login')
def equip_list(request):
    if request.method == 'GET':
        filtered_by = ''
        try:
            type = request.GET['type']
            frequency = request.GET['frequency']
            if type == 'all':
                objects_list = Type.objects.all()
            else:
                objects_list = Type.objects.filter(class_name=type)
                filtered_by = filtered_by + str(type)
            if frequency == 'all':
                pass
            else:
                objects_list = objects_list.filter(frequency_str=frequency)
                if filtered_by == '':
                    filtered_by = str(frequency)
                else:
                    filtered_by = filtered_by + ', ' + str(frequency)
        except KeyError:
            objects_list = Type.objects.all()
        filter_objects = Type.objects.all()
        context = {'objects': objects_list, 'filtered_by': filtered_by, 'filter_objects': filter_objects}
        return render(request, "rich/templates/equip_list.html", context)


@login_required(login_url='account:login')
@is_staff
def equip_create(request):
    form = TypeForm(request.POST, request.FILES or None)
    msg = ''
    if request.method == 'POST':
        if form.is_valid():
            f = form.save(commit=False)
            f.save()
            return redirect('rich:equip_list')
        else:
            msg = form_errors_text(form)
            context = {'msg': msg, 'form': form}
            return render(request, "rich/templates/equip_create.html", context)
    context = {'msg': msg, 'form': form}
    return render(request, "rich/templates/equip_create.html", context)


@login_required(login_url='account:login')
@is_staff
def equip_delete(request, key):
    obj = Type.objects.get(id=key)
    obj.delete()
    return redirect('rich:equip_list')


@login_required(login_url='account:login')
@is_staff
def change_equipment(request, key):
    res_object = Res.objects.get(id=key)
    if request.method == 'POST':
        try:
            equipment = request.POST['select_equipment']
            obj = Type.objects.get(name=equipment)
            res_object.type = obj
            res_object.save()
        except ObjectDoesNotExist:
            msg = 'Что то пошло не так!'
            res_detail(request, key, msg=msg)
    return redirect('rich:res_detail', id=key)


@login_required(login_url='account:login')
@is_staff
def add_equipment(request, key):
    res_object = Res.objects.get(id=key)
    form = TypeForm(request.POST, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            f = form.save(commit=False)
            f.save()
            res_object.type = f
            res_object.save()
            return redirect('rich:res_detail', id=key)
        else:
            msg = form_errors_text(form)
            print(msg)
            return res_detail(request, key, msg=msg)
    return redirect('rich:res_detail', id=key)


@login_required(login_url='account:login')
def protokol_list(request, key):
    res = Res.objects.get(id=key)
    list = ResProtokol.objects.filter(related_res=res)
    context = {'res': res, 'list': list}
    return render(request, "rich/templates/protokol_list.html", context)


@login_required(login_url='account:login')
@is_staff
def protokol_delete(request, key):
    obj = ResProtokol.objects.get(id=key)
    res_obj = obj.related_res.id
    obj.delete()
    return protokol_list(request, res_obj)


@login_required(login_url='account:login')
@is_staff
def add_res_protokol(request, key):
    form = ResProtokolForm(request.POST, request.FILES or None)
    res_obj = Res.objects.get(id=key)
    if request.method == 'POST':
        if form.is_valid():
            f = form.save(commit=False)
            f.related_res = res_obj
            f.save()
            return redirect('rich:res_detail', id=key)
        else:
            msg = form_errors_text(form)
            return res_detail(request, key, msg=msg)
    return redirect('rich:res_detail', id=key)