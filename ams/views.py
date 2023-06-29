from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.styles.numbers import BUILTIN_FORMATS

from ams.forms import AmsForm, AmsTypeForm, SchemeForm, FotoAmsForm, ProtocolForm, DiagnosticForm, ProjectForm, \
    SezForm, InventoryForm, PassportForm
from ams.models import Ams, Foto_Ams, Measurments, Diagnostic, Project, Sez, Inventory
from main.forms import FilterForm
from main.scripts import form_errors_text, object_filter, is_staff


@login_required(login_url='account:login')
def ams(request):
    filter_form = FilterForm()
    object_list = Ams.objects.all()
    filtered_by = []
    if request.method == 'GET':
        data = FilterForm(request.GET or None)
        objects, filtered_by = object_filter(data)
        object_list = object_list.filter(object_name__in=objects)
        try:
            min_visota = int(request.GET.get('min_visota'))
            max_visota = int(request.GET.get('max_visota'))
            otklonenie = request.GET.get('otklonenie')
            if min_visota >= max_visota or min_visota < 0 or max_visota < 0 or max_visota > 300 or min_visota == '' \
                    or max_visota == '':
                msg = 'Вы указали неверное значение высоты АМС'
                filtered_by_str = ' '.join(filtered_by)
                context = {'object': object_list, 'filter_form': data, 'filtered_by': filtered_by_str, 'msg': msg}
                return render(request, 'ams/templates/ams_list.html', context)
            elif min_visota == 0 and max_visota == 300:
                pass
            else:
                object_list = object_list.filter(height__gte=min_visota, height__lte=max_visota)
                filtered_by.append('Высоте')
            if otklonenie == '1':
                measurments = Measurments.objects.filter(is_otklonenie=True, is_last=True)
                object_list = object_list.filter(protocol__in=measurments)
                filtered_by.append('АМС с отклонением')
            elif otklonenie == '0':
                measurments = Measurments.objects.filter(is_otklonenie=False, is_last=True)
                object_list = object_list.filter(protocol__in=measurments)
                filtered_by.append('АМС без отклонения')
            elif otklonenie == '2':
                object_list = object_list.filter(protocol=None)
                filtered_by.append('АМС без измерений вертикальности')
        except (ValueError, TypeError):
            filtered_by_str = ' '.join(filtered_by)
            context = {'object': object_list, 'filter_form': data, 'filtered_by': filtered_by_str}
            return render(request, 'ams/templates/ams_list.html', context)
    filtered_by_str = ' '.join(filtered_by)
    context = {'object': object_list, 'filter_form': filter_form, 'filtered_by': filtered_by_str}
    return render(request, 'ams/templates/ams_list.html', context)


@login_required(login_url='account:login')
@is_staff
def ams_create(request):
    form = AmsForm(request.POST or None)
    msg = ''
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            return redirect('ams:ams')
        else:
            msg = form_errors_text(form)
    context = {'form': form, 'msg': msg}
    return render(request, 'ams/templates/ams_creation.html', context)


@login_required(login_url='account:login')
def export_xls_ams(request):
    object_list = Ams.objects.all()
    if request.method == 'GET':
        try:
            data = FilterForm(request.GET or None)
            objects, filtered_by = object_filter(data)
            object_list = object_list.filter(object_name__in=objects)
            min_visota = int(request.GET.get('min_visota'))
            max_visota = int(request.GET.get('max_visota'))
            otklonenie = request.GET.get('otklonenie')
            if min_visota >= max_visota or min_visota < 0 or max_visota < 0 or max_visota > 300 or min_visota == '' \
                    or max_visota == '':
                msg = 'Вы указали неверное значение высоты АМС'
                filtered_by_str = ' '.join(filtered_by)
                context = {'object': object_list, 'filter_form': data, 'filtered_by': filtered_by_str, 'msg': msg}
                return render(request, 'ams/templates/ams_list.html', context)
            elif min_visota == 0 and max_visota == 300:
                pass
            else:
                object_list = object_list.filter(height__gte=min_visota, height__lte=max_visota)
            if otklonenie == '1':
                measurments = Measurments.objects.filter(is_otklonenie=True, is_last=True)
                object_list = object_list.filter(protocol__in=measurments)
            elif otklonenie == '0':
                measurments = Measurments.objects.filter(is_otklonenie=False, is_last=True)
                object_list = object_list.filter(protocol__in=measurments)
            elif otklonenie == '2':
                object_list = object_list.filter(protocol=None)
        except (ValueError, TypeError):
            pass
        # выгрузка в xls -----------------------------------------------------
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="ams.xlsx"'
        wb = Workbook()
        rows = object_list
        ws = wb.active
        row_num = 1
        columns = ['Объект', 'Высота', 'Вес', 'Тип', 'Дата ввода в эксплуатацию', 'Количество ярусов оттяжек',
                   'Инвентарный номер', 'Владелец', 'Наименование ОС', 'Отклонение', 'Допуск']
        font = Font(name='TimesNewRoman', sz=11, bold=True)
        font2 = Font(name='TimesNewRoman', sz=11, bold=False)
        border = Border(left=Side(border_style='thin', color='FF000000'),
                        right=Side(border_style='thin', color='FF000000'),
                        top=Side(border_style='thin', color='FF000000'),
                        bottom=Side(border_style='thin', color='FF000000'))
        alignment = Alignment(horizontal='center',
                              vertical='center',
                              text_rotation=0,
                              wrap_text=True,
                              shrink_to_fit=False,
                              indent=0)
        for col_num in range(len(columns)):
            ws.cell(row_num, col_num + 1, columns[col_num])
            ws.cell(row_num, col_num + 1).font = font
            ws.cell(row_num, col_num + 1).border = border
            ws.cell(row_num, col_num + 1).alignment = alignment
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 20
        ws.column_dimensions['I'].width = 20
        ws.column_dimensions['J'].width = 20
        for row in range(rows.count()):
            atr_list = []
            if rows[row].inventory is None and rows[row].get_last_otklonenie() == '-':
                atr_list = [rows[row].object_name.object_name, rows[row].height, rows[row].weight, rows[row].type,
                            rows[row].date_birth, rows[row].otjazhki_count, '-', '-', '-', '-', rows[row].height]
            elif rows[row].inventory is None and rows[row].get_last_otklonenie() != '-':
                atr_list = [rows[row].object_name.object_name, rows[row].height, rows[row].weight, rows[row].type,
                            rows[row].date_birth, rows[row].otjazhki_count, '-', '-', '-',
                            rows[row].get_last_otklonenie().results, rows[row].height]
            elif rows[row].get_last_otklonenie() == '-' and rows[row].inventory is not None:
                atr_list = [rows[row].object_name.object_name, rows[row].height, rows[row].weight, rows[row].type,
                            rows[row].date_birth, rows[row].otjazhki_count, rows[row].inventory.inventory_number,
                            rows[row].inventory.owner, rows[row].inventory.inventory_name, '-', rows[row].height]
            else:
                atr_list = [rows[row].object_name.object_name, rows[row].height, rows[row].weight, rows[row].type,
                            rows[row].date_birth, rows[row].otjazhki_count, rows[row].inventory.inventory_number,
                            rows[row].inventory.owner, rows[row].inventory.inventory_name,
                            rows[row].get_last_otklonenie().results, rows[row].height]
            ws.cell(row + 2, 7).number_format = BUILTIN_FORMATS[1]
            for col_num in range(len(atr_list)):
                ws.cell(row + 2, col_num + 1, atr_list[col_num])
                ws.cell(row + 2, col_num + 1).font = font2
                ws.cell(row + 2, col_num + 1).border = border
                ws.cell(row + 2, col_num + 1).alignment = alignment
        wb.save(response)
        return response


@login_required(login_url='account:login')
@is_staff
def delete_ams(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        ams.delete()
        return redirect('ams:ams')
    except ObjectDoesNotExist:
        return redirect('ams:ams')


@login_required(login_url='account:login')
def ams_detail(request, ams_id, msg=''):
    type_form = AmsTypeForm()
    scheme_form = SchemeForm()
    foto_form = FotoAmsForm()
    diagnostic_form = DiagnosticForm()
    protocol_form = ProtocolForm()
    project_form = ProjectForm()
    sez_form = SezForm()
    inventory_form = InventoryForm()
    inventory = Inventory.objects.all()
    passport_form = PassportForm()
    try:
        ams_object = Ams.objects.get(id=ams_id)
        context = {'i': ams_object, 'msg': msg, 'type_form': type_form, 'scheme_form': scheme_form,
                   'foto_form': foto_form,
                   'protocol_form': protocol_form, 'diagnostic_form': diagnostic_form, 'project_form': project_form,
                   'sez_form': sez_form, 'inventory_form': inventory_form, 'inventory': inventory,
                   'passport_form': passport_form}
        return render(request, 'ams/templates/ams_details.html', context)
    except ObjectDoesNotExist:
        # msg = 'Такой АМС не существует'
        return redirect('ams:ams')


@login_required(login_url='account:login')
@is_staff
def change_height(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            height = int(request.POST['height'])
            ams.height = height
            ams.save()
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def change_weight(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            weight = int(request.POST['weight'])
            ams.weight = weight
            ams.save()
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def change_type(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            type = request.POST['type']
            ams.type = type
            ams.save()
            return redirect('ams:ams_detail', ams_id=ams_id)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def change_date(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            date = request.POST['date_birth']
            ams.date_birth = date
            ams.save()
            return redirect('ams:ams_detail', ams_id=ams_id)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def change_scheme(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = SchemeForm(request.POST, request.FILES)
            if form.is_valid():
                scheme = request.FILES['scheme']
                ams.scheme = scheme
                ams.save()
                return redirect('ams:ams_detail', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return ams_detail(request, ams_id, msg=msg)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def delete_scheme(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        ams.scheme = None
        ams.save()
        return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
def foto_ams_add(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = FotoAmsForm(request.POST, request.FILES)
            if form.is_valid():
                foto = request.FILES['foto']
                year = request.POST['year']
                foto_instance = ams
                foto_name = foto.name
                obj = Foto_Ams.objects.create(foto_instance=foto_instance, foto=foto, year=year, foto_name=foto_name)
                obj.save()
                return redirect('ams:ams_detail', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return ams_detail(request, ams_id, msg=msg)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def foto_delete(request, foto_id):
    try:
        foto = Foto_Ams.objects.get(id=foto_id)
        ams_id = foto.foto_instance.id
        foto.delete()
        return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams')


@login_required(login_url='account:login')
@is_staff
def add_protocol(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = ProtocolForm(request.POST, request.FILES)
            if form.is_valid():
                protocol_pdf = request.FILES['protocol_pdf']
                year = request.POST['year']
                measurment_instance = ams
                results = request.POST['results']
                obj = Measurments.objects.create(protocol_pdf=protocol_pdf, measurment_instance=measurment_instance,
                                                 year=year, results=results)
                obj.save()
                return redirect('ams:ams_detail', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return ams_detail(request, ams_id, msg=msg)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def add_diagnostic(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = DiagnosticForm(request.POST, request.FILES)
            if form.is_valid():
                result = request.FILES['result']
                year = request.POST['year']
                diagnostic_instance = ams
                obj = Diagnostic.objects.create(result=result, diagnostic_instance=diagnostic_instance,
                                                year=year)
                obj.save()
                return redirect('ams:ams_detail', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return ams_detail(request, ams_id, msg=msg)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def add_project(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = ProjectForm(request.POST, request.FILES)
            if form.is_valid():
                organization_name = request.POST['organization_name']
                documentation = request.FILES['documentation']
                year = request.POST['year']
                project_instance = ams
                obj = Project.objects.create(documentation=documentation, project_instance=project_instance,
                                             year=year, organization_name=organization_name)
                obj.save()
                return redirect('ams:ams_detail', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return redirect('ams:ams_detail', ams_id=ams_id, msg=msg)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def add_sez(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = SezForm(request.POST, request.FILES)
            if form.is_valid():
                sez_protocol = request.FILES['sez_protocol']
                year = request.POST['year']
                sez_instance = ams
                obj = Sez.objects.create(sez_protocol=sez_protocol, sez_instance=sez_instance, year=year)
                obj.save()
                return redirect('ams:ams_detail', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return ams_detail(request, ams_id, msg=msg)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def add_inventory(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = InventoryForm(request.POST)
            if form.is_valid():
                inventory_number = request.POST['inventory_number']
                inventory_name = request.POST['inventory_name']
                description = request.POST['description']
                owner = request.POST['owner']
                obj = Inventory.objects.create(inventory_number=inventory_number, inventory_name=inventory_name,
                                               description=description, owner=owner)
                obj.save()
                ams.inventory = obj
                ams.save()
                return redirect('ams:ams_detail', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return ams_detail(request, ams_id, msg=msg)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def select_inventory(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            select_inventory = request.POST['select_inventory']
            inventory = Inventory.objects.get(inventory_name=select_inventory)
            ams.inventory = inventory
            ams.save()
            return redirect('ams:ams_detail', ams_id=ams_id)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def add_passport(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = PassportForm(request.POST, request.FILES)
            if form.is_valid():
                passport = request.FILES['passport']
                ams.passport = passport
                ams.save()
                return redirect('ams:ams_detail', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return ams_detail(request, ams_id, msg=msg)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
@is_staff
def delete_passport(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            ams.passport = None
            ams.save()
            return redirect('ams:ams_detail', ams_id=ams_id)
        else:
            return redirect('ams:ams_detail', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)


@login_required(login_url='account:login')
def project_list(request, ams_id, msg=''):
    ams = Ams.objects.get(id=ams_id)
    project_list = Project.objects.filter(project_instance=ams)
    project_form = ProjectForm()
    context = {'ams': ams, 'project_list': project_list, 'project_form': project_form, 'msg': msg}
    return render(request, "ams/templates/project_list.html", context)


@login_required(login_url='account:login')
@is_staff
def project_delete(request, project_id):
    try:
        project = Project.objects.get(id=project_id)
        ams = project.project_instance
        project.delete()
        return redirect('ams:project_list', ams_id=ams.id)
    except ObjectDoesNotExist:
        return redirect('ams:ams')


@login_required(login_url='account:login')
@is_staff
def add_project_list(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = ProjectForm(request.POST, request.FILES)
            if form.is_valid():
                organization_name = request.POST['organization_name']
                documentation = request.FILES['documentation']
                year = request.POST['year']
                project_instance = ams
                obj = Project.objects.create(documentation=documentation, project_instance=project_instance,
                                             year=year, organization_name=organization_name)
                obj.save()
                return redirect('ams:project_list', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return redirect('ams:project_list', ams_id=ams_id, msg=msg)
        else:
            return redirect('ams:project_list', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams', ams_id=ams_id)


@login_required(login_url='account:login')
def do_list(request, ams_id, msg=''):
    ams = Ams.objects.get(id=ams_id)
    do_list = Diagnostic.objects.filter(diagnostic_instance=ams)
    diagnostic_form = DiagnosticForm()
    context = {'ams': ams, 'do_list': do_list, 'diagnostic_form': diagnostic_form, 'msg': msg}
    return render(request, "ams/templates/do_list.html", context)


@login_required(login_url='account:login')
@is_staff
def do_delete(request, do_id):
    try:
        do = Diagnostic.objects.get(id=do_id)
        ams = do.diagnostic_instance
        do.delete()
        return redirect('ams:do_list', ams_id=ams.id)
    except ObjectDoesNotExist:
        return redirect('ams:ams')


@login_required(login_url='account:login')
@is_staff
def add_do_list(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = DiagnosticForm(request.POST, request.FILES)
            if form.is_valid():
                result = request.FILES['result']
                year = request.POST['year']
                diagnostic_instance = ams
                obj = Diagnostic.objects.create(result=result, diagnostic_instance=diagnostic_instance,
                                                year=year)
                obj.save()
                return redirect('ams:do_list', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return redirect('ams:do_list', ams_id=ams_id, msg=msg)
        else:
            return redirect('ams:do_list', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams', ams_id=ams_id)


@login_required(login_url='account:login')
def measurement_list(request, ams_id, msg=''):
    ams = Ams.objects.get(id=ams_id)
    measurement_list = Measurments.objects.filter(measurment_instance=ams)
    measurement_form = ProtocolForm()
    context = {'ams': ams, 'measurement_list': measurement_list, 'measurement_form': measurement_form, 'msg': msg}
    return render(request, "ams/templates/measurement_list.html", context)


@login_required(login_url='account:login')
def measurement_delete(request, measurement_id):
    try:
        measurement = Measurments.objects.get(id=measurement_id)
        ams = measurement.measurment_instance
        measurement.delete()
        return redirect('ams:measurement_list', ams_id=ams.id)
    except ObjectDoesNotExist:
        return redirect('ams:ams')


@login_required(login_url='account:login')
@is_staff
def add_measurement_list(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = ProtocolForm(request.POST, request.FILES)
            if form.is_valid():
                protocol_pdf = request.FILES['protocol_pdf']
                year = request.POST['year']
                measurment_instance = ams
                results = request.POST['results']
                obj = Measurments.objects.create(protocol_pdf=protocol_pdf, measurment_instance=measurment_instance,
                                                 year=year, results=results)
                obj.save()
                return redirect('ams:measurement_list', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return redirect('ams:measurement_list', ams_id=ams_id, msg=msg)
        else:
            return redirect('ams:measurement_list', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams', ams_id=ams_id)


@login_required(login_url='account:login')
def sez_list(request, ams_id, msg=''):
    ams = Ams.objects.get(id=ams_id)
    sez_list = Sez.objects.filter(sez_instance=ams)
    sez_form = SezForm()
    context = {'ams': ams, 'sez_list': sez_list, 'sez_form': sez_form, 'msg': msg}
    return render(request, "ams/templates/sez_list.html", context)


@login_required(login_url='account:login')
@is_staff
def sez_delete(request, sez_id):
    try:
        sez = Sez.objects.get(id=sez_id)
        ams = sez.sez_instance
        sez.delete()
        return redirect('ams:sez_list', ams_id=ams.id)
    except ObjectDoesNotExist:
        return redirect('ams:ams')


@login_required(login_url='account:login')
@is_staff
def add_sez_list(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = SezForm(request.POST, request.FILES)
            if form.is_valid():
                sez_protocol = request.FILES['sez_protocol']
                year = request.POST['year']
                sez_instance = ams
                obj = Sez.objects.create(sez_protocol=sez_protocol, sez_instance=sez_instance, year=year)
                obj.save()
                return redirect('ams:sez_list', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return redirect('ams:sez_list', ams_id=ams_id, msg=msg)
        else:
            return redirect('ams:sez_list', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams', ams_id=ams_id)


@login_required(login_url='account:login')
def foto_list(request, ams_id, msg=''):
    ams = Ams.objects.get(id=ams_id)
    foto_list = Foto_Ams.objects.filter(foto_instance=ams)
    foto_form = FotoAmsForm()
    context = {'ams': ams, 'foto_list': foto_list, 'foto_form': foto_form, 'msg': msg}
    return render(request, "ams/templates/foto_list.html", context)


@login_required(login_url='account:login')
@is_staff
def foto_list_delete(request, foto_id):
    try:
        foto = Foto_Ams.objects.get(id=foto_id)
        ams = foto.foto_instance
        foto.delete()
        return redirect('ams:foto_list', ams_id=ams.id)
    except ObjectDoesNotExist:
        return redirect('ams:ams')


@login_required(login_url='account:login')
def add_foto_list(request, ams_id):
    try:
        ams = Ams.objects.get(id=ams_id)
        if request.method == 'POST':
            form = FotoAmsForm(request.POST, request.FILES)
            if form.is_valid():
                foto = request.FILES['foto']
                year = request.POST['year']
                foto_instance = ams
                foto_name = foto.name
                obj = Foto_Ams.objects.create(foto_instance=foto_instance, foto=foto, year=year, foto_name=foto_name)
                obj.save()
                return redirect('ams:foto_list', ams_id=ams_id)
            else:
                msg = form_errors_text(form)
                return redirect('ams:foto_list', ams_id=ams_id, msg=msg)
        else:
            return redirect('ams:foto_list', ams_id=ams_id)
    except ObjectDoesNotExist:
        return redirect('ams:ams_detail', ams_id=ams_id)
